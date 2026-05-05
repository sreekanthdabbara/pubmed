"""
PubMed Scraper - Multi-Keyword Search Version
Search multiple keywords and display results sorted by article count
"""

import os
import tempfile

# Load environment variables from .env file (local development)
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass  # dotenv not installed -- env vars must be set another way (e.g. Render dashboard)
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash
from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for, Response
from Bio import Entrez
import pandas as pd
import json
import io
import time
import re
import requests as http_requests

# ── OpenAI API helper ────────────────────────────────────────────────────────
def call_azure_openai(messages, max_tokens=1000, temperature=0.3):
    """Call OpenAI API (gpt-4o-mini). Returns (text, error_string)."""
    OPENAI_KEY   = os.environ.get('epilite-openai', '') or os.environ.get('OPENAI_API_KEY', '')
    OPENAI_MODEL = os.environ.get('OPENAI_MODEL', 'gpt-4o-mini')

    if not OPENAI_KEY:
        return None, 'OPENAI_API_KEY not set in Render environment variables.'

    try:
        resp = http_requests.post(
            'https://api.openai.com/v1/chat/completions',
            headers={
                'Content-Type':  'application/json',
                'Authorization': f'Bearer {OPENAI_KEY}',
            },
            json={
                'model':       OPENAI_MODEL,
                'messages':    messages,
                'max_tokens':  max_tokens,
                'temperature': temperature,
            },
            timeout=60,
        )
        if resp.status_code == 200:
            text = resp.json()['choices'][0]['message']['content']
            return text, None
        else:
            err = resp.json().get('error', {}).get('message', resp.text[:200])
            return None, f'OpenAI error {resp.status_code}: {err}'
    except Exception as e:
        return None, f'OpenAI exception: {e}'
from typing import List, Dict, Tuple
from datetime import datetime
from collections import OrderedDict
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading

try:
    import pdfplumber
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False
    print("⚠️  pdfplumber not installed -- PDF text extraction disabled.")
    print("   Run: pip install pdfplumber")

# ============================================================================
# MULTI-KEYWORD PUBMED SCRAPER
# ============================================================================

class MultiKeywordPubMedScraper:
    def __init__(self, email: str, api_key: str = None):
        """
        Initialize PubMed scraper.

        Args:
            email:   Your email (required by NCBI).
            api_key: Optional NCBI API key -- raises rate limit from 3 to 10
                     req/sec and speeds up large fetches significantly.
                     Get a free key at: https://www.ncbi.nlm.nih.gov/account/
        """
        Entrez.email = email
        if api_key:
            Entrez.api_key = api_key
        # Rate-limit: 3 req/s without key, 10 req/s with key
        self._min_interval = 0.11 if api_key else 0.34
        # Thread-safe lock so parallel workers don't race on Entrez calls
        self._lock = threading.Lock()

    def search_pubmed(self, query: str, max_results: int = 500) -> tuple:
        """Search PubMed and return (id_list, ncbi_total_count)."""
        print(f"  [search] {query!r} (max {max_results})")
        try:
            with self._lock:
                handle = Entrez.esearch(
                    db="pubmed",
                    term=query,
                    retmax=max_results,
                    sort="relevance",
                    usehistory="y",
                )
                results = Entrez.read(handle)
                handle.close()

            id_list    = results.get("IdList", [])
            ncbi_total = int(results.get("Count", len(id_list)))
            print(f"  [search] {len(id_list)} fetched / {ncbi_total} total for {query!r}")
            return id_list, ncbi_total

        except Exception as e:
            print(f"  [search] error for {query!r}: {e}")
            return [], 0

    def fetch_more(self, query: str, offset: int, batch: int = 500) -> List[Dict]:
        """Fetch the next batch of articles starting at offset. Used by Load More."""
        try:
            with self._lock:
                handle = Entrez.esearch(
                    db="pubmed",
                    term=query,
                    retmax=batch,
                    retstart=offset,
                    sort="relevance",
                )
                results = Entrez.read(handle)
                handle.close()
                time.sleep(self._min_interval)

            id_list = results.get("IdList", [])
            if not id_list:
                return []
            return self.fetch_abstracts(id_list)
        except Exception as e:
            print(f"  [fetch_more] error: {e}")
            return []

    def fetch_abstracts(self, pmid_list: List[str]) -> List[Dict]:
        """Fetch article details in large batches with minimal sleep."""
        if not pmid_list:
            return []

        articles = []
        # NCBI allows up to 10,000 per efetch; 500 is a safe sweet spot
        batch_size = 500

        for i in range(0, len(pmid_list), batch_size):
            batch = pmid_list[i:i + batch_size]
            try:
                with self._lock:
                    handle = Entrez.efetch(
                        db="pubmed",
                        id=batch,
                        rettype="abstract",
                        retmode="xml",
                    )
                    records = Entrez.read(handle)
                    handle.close()
                    time.sleep(self._min_interval)

                for record in records.get("PubmedArticle", []):
                    art = self._parse_article(record)
                    if art:
                        articles.append(art)

            except Exception as e:
                print(f"  [fetch] batch {i}-{i+batch_size} error: {e}")
                time.sleep(1)   # back off on error
                continue

        return articles
    
    def _parse_article(self, record) -> Dict:
        """Parse article record and extract relevant information"""
        try:
            article = record['MedlineCitation']['Article']
            pmid = str(record['MedlineCitation']['PMID'])
            
            title = article.get('ArticleTitle', 'N/A')
            
            abstract = 'N/A'
            if 'Abstract' in article:
                abstract_parts = article['Abstract'].get('AbstractText', [])
                if abstract_parts:
                    if isinstance(abstract_parts, list):
                        abstract = ' '.join([str(part) for part in abstract_parts])
                    else:
                        abstract = str(abstract_parts)
            
            # Extract authors and affiliations
            authors = []
            affiliations = []
            countries = set()  # Use set to avoid duplicates
            
            if 'AuthorList' in article:
                for author in article['AuthorList']:
                    if 'LastName' in author and 'ForeName' in author:
                        authors.append(f"{author['ForeName']} {author['LastName']}")
                    
                    # Extract affiliation information
                    if 'AffiliationInfo' in author:
                        for aff_info in author['AffiliationInfo']:
                            if 'Affiliation' in aff_info:
                                affiliation = aff_info['Affiliation']
                                affiliations.append(affiliation)
                                
                                # Extract country from affiliation
                                country = self._extract_country(affiliation)
                                if country:
                                    countries.add(country)
            
            # Join countries for display
            country_str = ', '.join(sorted(countries)) if countries else 'N/A'
            affiliation_str = '; '.join(affiliations[:3]) if affiliations else 'N/A'  # Limit to first 3
            
            journal = article.get('Journal', {}).get('Title', 'N/A')
            
            pub_date = 'N/A'
            if 'Journal' in article and 'JournalIssue' in article['Journal']:
                pub_date_info = article['Journal']['JournalIssue'].get('PubDate', {})
                year = pub_date_info.get('Year', '')
                month = pub_date_info.get('Month', '')
                pub_date = f"{month} {year}".strip()
            
            url = f"https://pubmed.ncbi.nlm.nih.gov/{pmid}/"

            # ── Publication types (e.g. Review, Clinical Trial) ──────────
            pub_types = []
            try:
                pt_list = article.get('PublicationTypeList', [])
                for pt in pt_list:
                    pt_str = str(pt).strip()
                    if pt_str and pt_str.lower() not in ('journal article',):
                        pub_types.append(pt_str)
                # If only "Journal Article" existed, keep it as fallback
                if not pub_types:
                    for pt in pt_list:
                        pub_types.append(str(pt).strip())
            except Exception:
                pass
            publication_types = '; '.join(pub_types) if pub_types else 'N/A'
            # Primary label = first meaningful type (shown as badge)
            pub_type_label = pub_types[0] if pub_types else 'N/A'
            # ────────────────────────────────────────────────────────────

            # ── PMC / Free full-text detection ──────────────────────────
            pmc_id = None
            is_free_pmc = False
            pmc_url = None
            pdf_url = None

            try:
                id_list = record.get('PubmedData', {}).get('ArticleIdList', [])
                for article_id in id_list:
                    if hasattr(article_id, 'attributes') and \
                            article_id.attributes.get('IdType') == 'pmc':
                        pmc_id = str(article_id)
                        break

                if pmc_id:
                    is_free_pmc = True
                    pmc_url = f"https://www.ncbi.nlm.nih.gov/pmc/articles/{pmc_id}/"
                    pdf_url = f"https://www.ncbi.nlm.nih.gov/pmc/articles/{pmc_id}/pdf/"
            except Exception:
                pass
            # ────────────────────────────────────────────────────────────

            return {
                'pmid': pmid,
                'title': title,
                'abstract': abstract,
                'authors': ', '.join(authors) if authors else 'N/A',
                'affiliation': affiliation_str,
                'country': country_str,
                'journal': journal,
                'publication_date': pub_date,
                'publication_type': publication_types,
                'pub_type_label': pub_type_label,
                'url': url,
                'pmc_id': pmc_id or 'N/A',
                'is_free_pmc': is_free_pmc,
                'pmc_url': pmc_url or '',
                'pdf_url': pdf_url or '',
            }
            
        except Exception as e:
            print(f"Error parsing article: {e}")
            return None
    
    def _extract_country(self, affiliation: str) -> str:
        """Extract country from affiliation string"""
        # Common country patterns in affiliations
        common_countries = {
            'USA': ['USA', 'United States', 'U.S.A', 'America'],
            'UK': ['UK', 'United Kingdom', 'England', 'Scotland', 'Wales'],
            'China': ['China', 'P.R. China', 'PR China'],
            'Germany': ['Germany'],
            'France': ['France'],
            'Japan': ['Japan'],
            'Canada': ['Canada'],
            'Australia': ['Australia'],
            'Italy': ['Italy'],
            'Spain': ['Spain'],
            'Netherlands': ['Netherlands', 'The Netherlands'],
            'Switzerland': ['Switzerland'],
            'Sweden': ['Sweden'],
            'India': ['India'],
            'Brazil': ['Brazil'],
            'South Korea': ['South Korea', 'Korea'],
            'Israel': ['Israel'],
            'Belgium': ['Belgium'],
            'Austria': ['Austria'],
            'Denmark': ['Denmark'],
            'Norway': ['Norway'],
            'Finland': ['Finland'],
            'Poland': ['Poland'],
            'Russia': ['Russia'],
            'Turkey': ['Turkey'],
            'Mexico': ['Mexico'],
            'Singapore': ['Singapore'],
            'Taiwan': ['Taiwan'],
            'Hong Kong': ['Hong Kong'],
            'New Zealand': ['New Zealand'],
            'Ireland': ['Ireland'],
            'Portugal': ['Portugal'],
        }
        
        affiliation_upper = affiliation.upper()
        
        for country, patterns in common_countries.items():
            for pattern in patterns:
                if pattern.upper() in affiliation_upper:
                    return country
        
        return None
    
    def _fetch_one_keyword(self, keyword: str, max_results: int) -> tuple:
        """Worker: search + fetch for a single keyword. Returns (keyword, df, ncbi_total)."""
        kw = keyword.strip()
        pmid_list, ncbi_total = self.search_pubmed(kw, max_results)
        if not pmid_list:
            return kw, pd.DataFrame(), ncbi_total
        articles = self.fetch_abstracts(pmid_list)
        articles = [a for a in articles if a is not None]
        if not articles:
            return kw, pd.DataFrame(), ncbi_total
        df = pd.DataFrame(articles)
        df["keyword"] = kw
        return kw, df, ncbi_total

    def search_multiple_keywords(
        self, keywords: List[str], max_results_per_keyword: int = 500
    ) -> Dict[str, pd.DataFrame]:
        """
        Search all keywords in parallel using a thread pool.

        Without an API key  → 3 workers (3 req/s NCBI limit).
        With an API key     → 5 workers (stays well under 10 req/s limit).
        """
        max_workers = 5 if getattr(Entrez, "api_key", None) else 3
        all_results: Dict[str, pd.DataFrame] = {}

        print(f"\nSearching {len(keywords)} keyword(s) with {max_workers} parallel workers...")

        # ncbi_totals stores the real NCBI count so results page can show Load More
        ncbi_totals: Dict[str, int] = {}

        with ThreadPoolExecutor(max_workers=max_workers) as pool:
            futures = {
                pool.submit(self._fetch_one_keyword, kw, max_results_per_keyword): kw
                for kw in keywords
            }
            for future in as_completed(futures):
                kw_label = futures[future]
                try:
                    kw, df, ncbi_total = future.result()
                    all_results[kw]  = df
                    ncbi_totals[kw]  = ncbi_total
                    print(f"  ✓ {kw!r}: {len(df)} fetched / {ncbi_total} total")
                except Exception as e:
                    print(f"  ✗ {kw_label!r}: {e}")
                    kw = kw_label.strip()
                    all_results[kw] = pd.DataFrame()
                    ncbi_totals[kw] = 0

        return all_results, ncbi_totals
    
    def sort_results_by_count(self, results: Dict[str, pd.DataFrame], ascending: bool = True) -> OrderedDict:
        """
        Sort results dictionary by number of articles (ascending or descending)
        
        Args:
            results: Dictionary of keyword -> DataFrame
            ascending: True for ascending order, False for descending
            
        Returns:
            OrderedDict sorted by article count
        """
        # Create list of (keyword, count, dataframe) tuples
        keyword_counts = [(kw, len(df), df) for kw, df in results.items()]
        
        # Sort by count
        keyword_counts.sort(key=lambda x: x[1], reverse=not ascending)
        
        # Convert back to OrderedDict
        sorted_results = OrderedDict()
        for kw, count, df in keyword_counts:
            sorted_results[kw] = df
        
        return sorted_results

    def _terms_from_keyword(self, keyword: str) -> List[str]:
        """Strip PubMed query syntax and return plain search terms."""
        cleaned = re.sub(r'\[\w+\]', '', keyword)
        cleaned = re.sub(r'\b(AND|OR|NOT)\b', ' ', cleaned, flags=re.IGNORECASE)
        cleaned = re.sub(r'[\"\(\)]', ' ', cleaned)
        cleaned = re.sub(r'\s+', ' ', cleaned).strip().lower()
        return [t for t in cleaned.split() if len(t) > 1]

    def compute_keyword_scores(
        self,
        all_results: Dict[str, pd.DataFrame],
        keywords: List[str],
    ) -> Dict[str, pd.DataFrame]:
        """
        Vectorised scoring -- runs in milliseconds instead of minutes.

        keyword_match_count  -- how many searched keywords this PMID appeared in
        keyword_total_hits   -- total plain-term occurrences in title + abstract
        """
        # ── PMID → set of matched keywords ─────────────────────────────
        # Tells us exactly which keywords each article appeared in
        pmid_kw_set: Dict[str, set] = {}
        for kw, df in all_results.items():
            if df.empty:
                continue
            for pmid in df["pmid"].astype(str):
                pmid_kw_set.setdefault(pmid, set()).add(kw)

        # ── Per-keyword term patterns (for individual hit counts) ────────
        # kw_patterns: keyword -> compiled regex of its plain terms
        kw_patterns: Dict[str, re.Pattern] = {}
        for kw in keywords:
            terms = self._terms_from_keyword(kw)
            if terms:
                kw_patterns[kw] = re.compile(
                    "|".join(re.escape(t) for t in terms), re.IGNORECASE
                )

        # Combined pattern for total hits across ALL keywords
        all_terms = list({t for terms in
                          [self._terms_from_keyword(kw) for kw in keywords]
                          for t in terms})
        combined_pattern = re.compile(
            "|".join(re.escape(t) for t in all_terms), re.IGNORECASE
        ) if all_terms else None

        # ── Annotate each DataFrame ──────────────────────────────────────
        updated: Dict[str, pd.DataFrame] = {}
        for kw, df in all_results.items():
            if df.empty:
                updated[kw] = df
                continue

            df = df.copy()
            pmids = df["pmid"].astype(str)

            # keyword_match_count -- how many searched keywords this article appeared in
            df["keyword_match_count"] = pmids.map(
                lambda p: len(pmid_kw_set.get(p, {kw}))
            ).astype(int)

            # matched_keywords -- the actual keyword names (semicolon-separated)
            df["matched_keywords"] = pmids.map(
                lambda p: "; ".join(sorted(pmid_kw_set.get(p, {kw})))
            )

            # keyword_total_hits -- total term occurrences across all keywords
            text_series = (
                df["title"].fillna("") + " " + df["abstract"].fillna("")
            ).str.lower()

            if combined_pattern:
                df["keyword_total_hits"] = text_series.apply(
                    lambda t: len(combined_pattern.findall(t))
                )
            else:
                df["keyword_total_hits"] = 0

            # per_keyword_hits -- "lung cancer:5; breast cancer:2" style breakdown
            def per_kw_hits(text):
                parts = []
                for k, pat in kw_patterns.items():
                    count = len(pat.findall(text))
                    if count > 0:
                        parts.append(f"{k}: {count}")
                return "; ".join(parts) if parts else ""

            df["per_keyword_hits"] = text_series.apply(per_kw_hits)

            updated[kw] = df

        return updated


# ============================================================================
# PDF TEXT EXTRACTION HELPER
# ============================================================================

EXCEL_CELL_LIMIT = 32_000  # Excel hard limit is 32,767 chars per cell


def fetch_pdf_text(pdf_url: str, timeout_secs: int = 20) -> str:
    """
    Fetch full text for a free PMC article using the official PMC XML API
    (Entrez efetch db=pmc). Strips the References section and returns
    complete body text in document order, preserving all paragraphs.
    """
    if not pdf_url:
        return ''

    # ── Extract PMC ID from the URL ──────────────────────────────────────
    pmc_match = re.search(r'PMC(\d+)', pdf_url, re.IGNORECASE)
    if not pmc_match:
        return 'Could not extract PMC ID from URL'

    pmc_numeric_id = pmc_match.group(1)

    try:
        import socket
        socket.setdefaulttimeout(timeout_secs)
        # ── Fetch full-text XML via Entrez ────────────────────────────────
        handle = Entrez.efetch(
            db='pmc',
            id=pmc_numeric_id,
            rettype='full',
            retmode='xml',
        )
        xml_bytes = handle.read()
        handle.close()

        # ── Parse XML ─────────────────────────────────────────────────────
        from xml.etree import ElementTree as ET
        try:
            root = ET.fromstring(xml_bytes)
        except ET.ParseError:
            return 'Could not parse PMC XML response'

        # Tags whose entire subtree content should be skipped
        # (but NOT their .tail -- tail belongs to the parent level)
        SKIP_CONTENT = {
            'ref-list',    # References list
            'fn-group',    # Footnotes
            'glossary',    # Glossary
            'ack',         # Acknowledgements
            'xref',        # Inline citation numbers  e.g. [1]
            'ext-link',    # External hyperlinks
            'uri',         # URLs
            'graphic',     # Image placeholders
            'media',       # Media elements
            'supplementary-material',
        }

        # Tags that represent block-level breaks (add newline before/after)
        BLOCK_TAGS = {
            'p', 'title', 'sec', 'list-item', 'caption',
            'table-wrap', 'fig', 'disp-formula', 'def-item',
        }

        def local(tag):
            """Strip namespace prefix from tag name."""
            return tag.split('}')[-1] if '}' in tag else tag

        def walk(el):
            """
            Recursively collect text in document order.

            Returns a string that preserves:
              el.text  -- text immediately inside the opening tag
              children -- each child's full text contribution
              el.tail  -- text immediately after the closing tag (parent context)

            When el is in SKIP_CONTENT its internal text and children are
            suppressed, but el.tail is still returned (it belongs to the
            parent paragraph, not to the skipped element).
            """
            tag = local(el.tag)

            # Skipped element: return only tail so parent paragraph stays intact
            if tag in SKIP_CONTENT:
                return el.tail or ''

            parts = []

            # Block elements get a leading newline so paragraphs don't run together
            if tag in BLOCK_TAGS:
                parts.append('\n')

            # Text directly inside this element (before any child)
            if el.text:
                parts.append(el.text)

            # Recurse into every child; walk() returns child text + child.tail
            for child in el:
                parts.append(walk(child))

            # Trailing newline for block elements
            if tag in BLOCK_TAGS:
                parts.append('\n')

            # Tail: text after this element's closing tag (belongs to parent)
            if el.tail:
                parts.append(el.tail)

            return ''.join(parts)

        # ── Find body (try JATS namespace first, then bare) ───────────────
        ns = 'http://jats.nlm.nih.gov'
        body = (root.find(f'{{{ns}}}body') or
                root.find(f'.//{{{ns}}}body') or
                root.find('.//body'))

        if body is not None:
            raw = walk(body)
        else:
            # Fallback: whole document minus front/back matter
            raw = walk(root)

        # ── Clean up whitespace ───────────────────────────────────────────
        # Collapse runs of spaces (but keep newlines)
        lines = []
        for line in raw.splitlines():
            line = re.sub(r'[ \t]+', ' ', line).strip()
            lines.append(line)

        # Collapse 3+ consecutive blank lines into 2
        full_text = re.sub(r'\n{3,}', '\n\n', '\n'.join(lines)).strip()

        if not full_text:
            return 'No text could be extracted for this article'

        # ── Strip References section if it survived ───────────────────────
        ref_pat = re.compile(
            r'\n\s*(?:REFERENCES?|Reference List|BIBLIOGRAPHY|Bibliography)\s*\n',
            re.IGNORECASE
        )
        m = ref_pat.search(full_text)
        if m:
            full_text = full_text[:m.start()].rstrip()

        # Return full text -- truncation for Excel is applied in the Excel
        # writer path only so CSV and JSON always get the complete content
        return full_text

    except Exception as e:
        return f'Error fetching full text: {str(e)}'


# ── Desired column order in all exports ─────────────────────────────────────
EXPORT_COLUMNS = [
    'keyword',
    'pmid',
    'title',
    'publication_type',
    'authors',
    'affiliation',
    'country',
    'journal',
    'publication_date',
    'abstract',
    'pmc_id',
    'url',
    'keyword_match_count',
    'matched_keywords',
    'per_keyword_hits',
    'keyword_total_hits',
    'pdf_total_keyword_hits',          # keyword hit count in PDF full text
    'pdf_per_keyword_hits',            # per-keyword breakdown in PDF
    'Free article complete content',   # LAST -- full text (may span Part 2, Part 3 cols)
]

# Columns that should never appear in exports
EXPORT_DROP = {'pub_type_label', 'pdf_url', 'pmc_url', 'is_free_pmc'}

# ── Column sets per export mode ───────────────────────────────────────────────
# Mode: "abstract"  → no full text, no PDF scores
ABSTRACT_ONLY_COLUMNS = [
    'keyword', 'pmid', 'title', 'publication_type', 'authors',
    'affiliation', 'country', 'journal', 'publication_date',
    'abstract', 'pmc_id', 'url',
    'keyword_match_count', 'matched_keywords',
    'per_keyword_hits', 'keyword_total_hits',
]

# Mode: "pmc_only"  → same as abstract but filtered to free PMC articles
# (column set same as ABSTRACT_ONLY_COLUMNS -- filter applied on rows)

# Mode: default / "full"  → full export including full text and PDF scores
# (uses full EXPORT_COLUMNS list already defined above)


def clean_for_export(df: pd.DataFrame) -> pd.DataFrame:
    """Drop internal columns, enforce friendly column order, rename headers."""
    df = df.copy()

    # Drop unwanted columns
    drop = [c for c in EXPORT_DROP if c in df.columns]
    if drop:
        df = df.drop(columns=drop)

    # Reorder: put known columns first, then anything extra
    ordered = [c for c in EXPORT_COLUMNS if c in df.columns]
    extras  = [c for c in df.columns if c not in set(EXPORT_COLUMNS)]
    df = df[ordered + extras]

    # Friendly header rename
    rename_map = {
        'pmid':                          'PMID',
        'title':                         'Title',
        'publication_type':              'Publication Type',
        'authors':                       'Authors',
        'affiliation':                   'Affiliation',
        'country':                       'Country',
        'journal':                       'Journal',
        'publication_date':              'Publication Date',
        'abstract':                      'Abstract',
        'pmc_id':                        'PMC ID',
        'url':                           'PubMed URL',
        'keyword':                       'Keyword',
        'keyword_match_count':           'Keyword Match Count',
        'matched_keywords':              'Matched Keywords',
        'per_keyword_hits':              'Hits Per Keyword',
        'keyword_total_hits':            'Total Keyword Hits',
        'Free article complete content': 'Full Text (PMC)',
        'pdf_total_keyword_hits':        'PDF Total Keyword Hits',
        'pdf_per_keyword_hits':          'PDF Hits Per Keyword',
    }
    df = df.rename(columns={k: v for k, v in rename_map.items() if k in df.columns})
    return df


def add_full_text_column(df, keywords=None):
    """
    Add full text column plus PDF keyword hit counts for each free PMC article.

    keywords: list of original search keywords -- used to count hits in the PDF.
              If None, hit columns are skipped.
    """
    df = df.copy()
    content_col        = []
    pdf_total_col      = []
    pdf_per_kw_col     = []

    # Pre-compile per-keyword patterns once (reuse scraper helper via inline logic)
    kw_patterns = {}
    if keywords:
        for kw in keywords:
            # Strip PubMed syntax the same way _terms_from_keyword does
            cleaned = re.sub(r'\[\w+\]', '', kw)
            cleaned = re.sub(r'\b(AND|OR|NOT)\b', ' ', cleaned, flags=re.IGNORECASE)
            cleaned = re.sub(r'["\(\)]', ' ', cleaned)
            cleaned = re.sub(r'\s+', ' ', cleaned).strip().lower()
            terms   = [t for t in cleaned.split() if len(t) > 1]
            if terms:
                kw_patterns[kw] = re.compile(
                    '|'.join(re.escape(t) for t in terms), re.IGNORECASE
                )

    for _, row in df.iterrows():
        if row.get('is_free_pmc') and row.get('pdf_url'):
            try:
                print(f"  Fetching PDF for PMID {row.get('pmid', '?')} ...")
                text = fetch_pdf_text(row['pdf_url'])
            except Exception as e:
                print(f"  Full text fetch failed: {e}")
                text = ''
        else:
            text = ''

        content_col.append(text)

        # Count keyword hits in full text
        if text and kw_patterns:
            text_lower  = text.lower()
            total_hits  = 0
            per_kw_parts = []
            for kw, pat in kw_patterns.items():
                count = len(pat.findall(text_lower))
                total_hits += count
                if count > 0:
                    per_kw_parts.append(f"{kw}: {count}")
            pdf_total_col.append(total_hits)
            pdf_per_kw_col.append('; '.join(per_kw_parts) if per_kw_parts else '')
        else:
            pdf_total_col.append(0)
            pdf_per_kw_col.append('')

    df['Free article complete content'] = content_col
    if keywords:
        df['pdf_total_keyword_hits'] = pdf_total_col
        df['pdf_per_keyword_hits']   = pdf_per_kw_col

    # Drop internal-only columns not needed in exports
    drop_cols = [c for c in ['pub_type_label'] if c in df.columns]
    if drop_cols:
        df = df.drop(columns=drop_cols)

    return df

# ============================================================================
# FLASK APPLICATION
# ============================================================================

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('FLASK_SECRET', 'change-me-in-production')

# ── User accounts ─────────────────────────────────────────────────────────────
# Simple in-memory user store. For production use a database.
# HOW TO ADD A USER:
#   from werkzeug.security import generate_password_hash
#   print(generate_password_hash("their_password"))
#   # paste the output as the password value below

USERS = {
    # email: hashed_password
    # Add your users here. Generate hash with: generate_password_hash("password")
    "admin@episcience.com": generate_password_hash("admin123"),
    "sreekanth.dabbara@gmail.com": generate_password_hash("epi2024"),
}

def login_required(f):
    """Decorator -- redirects to login page if user is not logged in."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_email' not in session:
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated

# ── NCBI credentials ──────────────────────────────────────────────────────────
# REQUIRED: change to your email
NCBI_EMAIL = "your.email@example.com"

# OPTIONAL but RECOMMENDED: free API key from https://www.ncbi.nlm.nih.gov/account/
# Adding a key raises the rate limit from 3 → 10 req/s and cuts search time ~3x
NCBI_API_KEY = None   # e.g. "abc123def456..."

# Default max articles fetched per keyword (user can override on results page)
DEFAULT_MAX_RESULTS = 500

scraper = MultiKeywordPubMedScraper(email=NCBI_EMAIL, api_key=NCBI_API_KEY)

recent_searches = []

# In-memory cache: search_id -> {keyword_summary, sorted_results_records}
# Keyed by a short hash so exports read the same data the user sees
import hashlib
_search_cache: dict = {}
CACHE_MAX = 20   # keep last 20 searches in memory


# ── Auth routes ──────────────────────────────────────────────────────────────

@app.route('/login', methods=['GET', 'POST'])
def login_page():
    """Login page."""
    if 'user_email' in session:
        return redirect(url_for('index'))   # already logged in

    if request.method == 'POST':
        email    = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '')
        remember = request.form.get('remember')

        user_hash = USERS.get(email)
        if user_hash and check_password_hash(user_hash, password):
            session.permanent = bool(remember)
            session['user_email'] = email
            session['user_name']  = email.split('@')[0].title()
            return redirect(url_for('index'))
        else:
            return render_template('login.html',
                                   error="Invalid email or password. Please try again.",
                                   email=email)

    return render_template('login.html')


@app.route('/logout')
def logout():
    """Clear session and redirect to login."""
    session.clear()
    return redirect(url_for('login_page'))


@app.route('/register', methods=['GET', 'POST'])
def register():
    """Simple self-registration -- adds user to in-memory store."""
    if request.method == 'POST':
        email    = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '')
        confirm  = request.form.get('confirm_password', '')

        if not email or not password:
            return render_template('register.html', error="Email and password are required.")
        if password != confirm:
            return render_template('register.html', error="Passwords do not match.", email=email)
        if len(password) < 6:
            return render_template('register.html', error="Password must be at least 6 characters.", email=email)
        if email in USERS:
            return render_template('register.html', error="An account with that email already exists.", email=email)

        USERS[email] = generate_password_hash(password)
        return render_template('login.html', message="Account created! You can now log in.")

    return render_template('register.html')


@app.route('/forgot-password')
def forgot_password():
    """Placeholder -- implement email reset for production."""
    return render_template('login.html',
                           message="Password reset is not yet configured. Contact your administrator.")


# ── Main app routes (all protected by login_required) ────────────────────────

@app.route('/')
@login_required
def index():
    """Home page with search form"""
    return render_template('index_multi_with_logo.html', recent_searches=recent_searches[:10])


@app.route('/search', methods=['POST'])
@login_required
def search():
    """Handle multi-keyword search requests -- fetches ALL results, filtering done on results page"""
    try:
        keywords_input = request.form.get('keywords', '').strip()

        if not keywords_input:
            return render_template('index_multi_with_logo.html',
                                 error="Please enter at least one keyword",
                                 recent_searches=recent_searches[:10])

        # Parse keywords -- comma-separated or one per line
        if ',' in keywords_input:
            keywords = [k.strip() for k in keywords_input.split(',') if k.strip()]
        else:
            keywords = [k.strip() for k in keywords_input.split('\n') if k.strip()]

        if not keywords:
            return render_template('index_multi_with_logo.html',
                                 error="No valid keywords found",
                                 recent_searches=recent_searches[:10])

        print(f"\nSearching for {len(keywords)} keywords (all results): {keywords}")

        # Respect user-selected fetch limit (defaults to DEFAULT_MAX_RESULTS)
        max_results = int(request.form.get('max_results', DEFAULT_MAX_RESULTS))
        max_results = max(1, min(max_results, 10000))  # clamp to safe range

        # Fetch up to max_results per keyword; user can limit display on results page
        all_results, ncbi_totals = scraper.search_multiple_keywords(keywords, max_results_per_keyword=max_results)

        # Compute cross-keyword match scores for every article
        all_results = scraper.compute_keyword_scores(all_results, keywords)

        # Default initial sort: ascending by article count
        sorted_results = scraper.sort_results_by_count(all_results, ascending=True)

        # Prepare full dataset for template
        keyword_summary = []
        total_articles = 0

        for keyword, df in sorted_results.items():
            count = len(df)
            total_articles += count

            # Pass records unsorted; JS handles all article-level sorting
            records = df.to_dict('records') if not df.empty else []
            free_count = sum(1 for r in records if r.get('is_free_pmc'))
            ncbi_total = ncbi_totals.get(keyword, count)

            keyword_summary.append({
                'keyword': keyword,
                'count': count,
                'ncbi_total': ncbi_total,          # real NCBI total for Load More
                'fetched': count,                   # how many we actually fetched
                'has_more': ncbi_total > count,     # whether Load More button shows
                'free_count': free_count,
                'results': records
            })

        # Store recent search record
        search_record = {
            'query': f"{len(keywords)} keyword(s): {', '.join(keywords[:3])}{'...' if len(keywords) > 3 else ''}",
            'count': total_articles,
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        recent_searches.insert(0, search_record)
        if len(recent_searches) > 50:
            recent_searches.pop()

        total_free_articles  = sum(s['free_count'] for s in keyword_summary)
        total_ncbi_articles  = sum(s['ncbi_total'] for s in keyword_summary)
        keywords_str = ','.join(keywords)

        # ── Cache results so export doesn't need to re-search ────────────
        search_id = hashlib.md5(
            (keywords_str + str(max_results)).encode()
        ).hexdigest()[:12]
        _search_cache[search_id] = {
            'sorted_results': {kw: df for kw, df in sorted_results.items()},
            'keywords':       keywords,
            'sort_order':     'ascending',
            'max_results':    max_results,
        }
        # Keep cache bounded
        if len(_search_cache) > CACHE_MAX:
            oldest = next(iter(_search_cache))
            del _search_cache[oldest]
        # ─────────────────────────────────────────────────────────────────

        return render_template('results_multi.html',
                             keywords=keywords,
                             keywords_str=keywords_str,
                             search_id=search_id,
                             keyword_summary=keyword_summary,
                             total_articles=total_articles,
                             total_ncbi_articles=total_ncbi_articles,
                             total_free_articles=total_free_articles)

    except Exception as e:
        return render_template('index_multi_with_logo.html',
                             error=f"An error occurred: {str(e)}",
                             recent_searches=recent_searches[:10])


@app.route('/search_url', methods=['POST'])
@login_required
def search_url():
    """
    Handle PubMed URL-based search.
    Parses the URL, extracts the term + filters, converts them to a valid
    Entrez query, then runs the same pipeline as /search.
    """
    from urllib.parse import urlparse, parse_qs

    # ── PubMed filter code → Entrez query clause ─────────────────────────
    FILTER_MAP = {
        # Date filters
        'datesearch.y_1':              '("last 1 year"[PDat])',
        'datesearch.y_5':              '("last 5 years"[PDat])',
        'datesearch.y_10':             '("last 10 years"[PDat])',
        'datesearch.y_20':             '("last 20 years"[PDat])',
        # Language
        'lang.english':                '"english"[Language]',
        'lang.french':                 '"french"[Language]',
        'lang.german':                 '"german"[Language]',
        'lang.spanish':                '"spanish"[Language]',
        'lang.italian':                '"italian"[Language]',
        'lang.portuguese':             '"portuguese"[Language]',
        'lang.chinese':                '"chinese"[Language]',
        'lang.japanese':               '"japanese"[Language]',
        # Species / Age / Sex (filter= format)
        'hum_ani.humans':              '"humans"[MeSH Terms]',
        'hum_ani.animals':             '"animals"[MeSH Terms]',
        'ages.child':                  '"child"[MeSH Terms]',
        'ages.infant':                 '"infant"[MeSH Terms]',
        'ages.adult':                  '"adult"[MeSH Terms]',
        'ages.aged':                   '"aged"[MeSH Terms]',
        'ffrft.Y':                     '"full text"[Filter]',
        'simsearch2.ffrft':            '"free full text"[Filter]',
        # Article types — type.* format (older PubMed URLs)
        'type.clinicaltrial':          '"Clinical Trial"[Publication Type]',
        'type.review':                 '"Review"[Publication Type]',
        'type.systematicreview':       '"Systematic Review"[Publication Type]',
        'type.randomizedcontrolledtrial': '"Randomized Controlled Trial"[Publication Type]',
        'type.metaanalysis':           '"Meta-Analysis"[Publication Type]',
        'type.casereports':            '"Case Reports"[Publication Type]',
        'type.observationalstudy':     '"Observational Study"[Publication Type]',
        'type.journal':                '"Journal Article"[Publication Type]',
        'sex.female':                  '"female"[MeSH Terms]',
        'sex.male':                    '"male"[MeSH Terms]',
        # Article types — pubt.* format (current PubMed URLs)
        'pubt.adaptiveclinicaltrial':         '"Adaptive Clinical Trial"[Publication Type]',
        'pubt.address':                       '"Address"[Publication Type]',
        'pubt.biography':                     '"Biography"[Publication Type]',
        'pubt.booksanddocuments':             '"Books and Documents"[Publication Type]',
        'pubt.casereports':                   '"Case Reports"[Publication Type]',
        'pubt.clinicalstudy':                 '"Clinical Study"[Publication Type]',
        'pubt.clinicaltrial':                 '"Clinical Trial"[Publication Type]',
        'pubt.clinicaltrialprotocol':         '"Clinical Trial Protocol"[Publication Type]',
        'pubt.clinicaltrialphase1':           '"Clinical Trial, Phase I"[Publication Type]',
        'pubt.clinicaltrialphase2':           '"Clinical Trial, Phase II"[Publication Type]',
        'pubt.clinicaltrialphase3':           '"Clinical Trial, Phase III"[Publication Type]',
        'pubt.clinicaltrialphase4':           '"Clinical Trial, Phase IV"[Publication Type]',
        'pubt.clinicaltrialphaseii':          '"Clinical Trial, Phase II"[Publication Type]',
        'pubt.clinicaltrialphaseiii':         '"Clinical Trial, Phase III"[Publication Type]',
        'pubt.clinicaltrialphaseiv':          '"Clinical Trial, Phase IV"[Publication Type]',
        'pubt.clinicaltrial,veterinary':      '"Clinical Trial, Veterinary"[Publication Type]',
        'pubt.collectedwork':                 '"Collected Work"[Publication Type]',
        'pubt.comment':                       '"Comment"[Publication Type]',
        'pubt.comparativestudy':              '"Comparative Study"[Publication Type]',
        'pubt.conferenceproceedings':         '"Conference Proceedings"[Publication Type]',
        'pubt.consensusstatement':            '"Consensus Development Conference"[Publication Type]',
        'pubt.controlledclinicaltrial':       '"Controlled Clinical Trial"[Publication Type]',
        'pubt.dataset':                       '"Dataset"[Publication Type]',
        'pubt.editorial':                     '"Editorial"[Publication Type]',
        'pubt.equivalencetrial':              '"Equivalence Trial"[Publication Type]',
        'pubt.evaluationstudy':               '"Evaluation Study"[Publication Type]',
        'pubt.guideline':                     '"Guideline"[Publication Type]',
        'pubt.historicalarticle':             '"Historical Article"[Publication Type]',
        'pubt.interview':                     '"Interview"[Publication Type]',
        'pubt.letter':                        '"Letter"[Publication Type]',
        'pubt.meta-analysis':                 '"Meta-Analysis"[Publication Type]',
        'pubt.multicenterstudy':              '"Multicenter Study"[Publication Type]',
        'pubt.news':                          '"News"[Publication Type]',
        'pubt.observationalstudy':            '"Observational Study"[Publication Type]',
        'pubt.practiceguideline':             '"Practice Guideline"[Publication Type]',
        'pubt.pragmaticclinicaltrial':        '"Pragmatic Clinical Trial"[Publication Type]',
        'pubt.preprint':                      '"Preprint"[Publication Type]',
        'pubt.randomizedcontrolledtrial':     '"Randomized Controlled Trial"[Publication Type]',
        'pubt.review':                        '"Review"[Publication Type]',
        'pubt.scopingreview':                 '"Systematic Review"[Publication Type]',
        'pubt.systematicreview':              '"Systematic Review"[Publication Type]',
        'pubt.twinstudy':                     '"Twin Study"[Publication Type]',
        'pubt.validationstudy':               '"Validation Study"[Publication Type]',
        # Has abstract / full text
        'simsearch2.ffrft':                   '"free full text"[Filter]',
        'ffrft.Y':                            '"full text"[Filter]',
        # Species / Age / Sex
        'hum_ani.humans':                     '"humans"[MeSH Terms]',
        'hum_ani.animals':                    '"animals"[MeSH Terms]',
        'sex.female':                         '"female"[MeSH Terms]',
        'sex.male':                           '"male"[MeSH Terms]',
        'ages.child':                         '"child"[MeSH Terms]',
        'ages.adult':                         '"adult"[MeSH Terms]',
        'ages.aged':                          '"aged"[MeSH Terms]',
        'ages.infant':                        '"infant"[MeSH Terms]',
    }

    try:
        pubmed_url  = request.form.get('pubmed_url', '').strip()
        max_results = int(request.form.get('max_results', DEFAULT_MAX_RESULTS))
        max_results = max(1, min(max_results, 10000))

        if not pubmed_url:
            return render_template('index_multi_with_logo.html',
                                 error="Please enter a PubMed URL",
                                 recent_searches=recent_searches[:10])

        # ── Parse URL ────────────────────────────────────────────────────
        parsed = urlparse(pubmed_url)

        # Accept both pubmed.ncbi.nlm.nih.gov and ncbi.nlm.nih.gov/pubmed
        if 'pubmed' not in parsed.netloc + parsed.path:
            return render_template('index_multi_with_logo.html',
                                 error="Please enter a valid PubMed URL "
                                       "(must contain pubmed.ncbi.nlm.nih.gov)",
                                 recent_searches=recent_searches[:10])

        params  = parse_qs(parsed.query, keep_blank_values=False)
        term    = params.get('term', [''])[0].strip()
        filters = params.get('filter', [])
        sort    = params.get('sort', ['relevance'])[0]
        if not term:
            return render_template('index_multi_with_logo.html',
                                 error="No search term found in the URL. "
                                       "Make sure the URL includes a ?term= parameter.",
                                 recent_searches=recent_searches[:10])

        # ── Convert filters to Entrez clauses ────────────────────────────
        pub_type_clauses = []   # joined with OR  (any type matches)
        other_clauses    = []   # joined with AND (each is a separate constraint)
        unknown_filters  = []
        seen_clauses     = set()

        for f in filters:
            # ── Special case: years.YYYY-YYYY date range ──────────────────
            if f.startswith('years.'):
                try:
                    year_part = f.replace('years.', '')
                    y_from, y_to = year_part.split('-')
                    clause = f'("{y_from}/01/01"[PDat] : "{y_to}/12/31"[PDat])'
                    if clause not in seen_clauses:
                        seen_clauses.add(clause)
                        other_clauses.append(clause)
                except Exception:
                    unknown_filters.append(f)
                continue

            # ── Special case: datesearch.y_N relative date ────────────────
            if f.startswith('datesearch.y_'):
                try:
                    years_back = int(f.replace('datesearch.y_', ''))
                    from datetime import date
                    y_from = date.today().year - years_back
                    clause = f'("{y_from}/01/01"[PDat] : "3000"[PDat])'
                    if clause not in seen_clauses:
                        seen_clauses.add(clause)
                        other_clauses.append(clause)
                except Exception:
                    unknown_filters.append(f)
                continue

            # ── Special case: englishabstract — add to OR pub type group ──
            # PubMed's englishabstract means articles with English abstract
            # Adding as AND "english"[Language] is too restrictive (643 vs 884)
            # Treat as pub type in OR group to match PubMed behavior
            if f == 'pubt.englishabstract':
                clause = '"English Abstract"[Publication Type]'
                if clause not in seen_clauses:
                    seen_clauses.add(clause)
                    pub_type_clauses.append(clause)
                continue

            # ── Special case: simsearch1.fha (has abstract) ───────────────
            # PubMed web UI filter — drop it entirely, Entrez handles abstracts differently
            # The 126 vs 109 difference comes from this filter being too restrictive
            if f == 'simsearch1.fha':
                # Don't add to query — Entrez already returns articles with abstracts
                # when they have them; forcing this filter removes some valid results
                continue

            if f in FILTER_MAP:
                clause = FILTER_MAP[f]
                if clause in seen_clauses:
                    continue
                seen_clauses.add(clause)
                if '[Publication Type]' in clause:
                    pub_type_clauses.append(clause)
                else:
                    other_clauses.append(clause)
            else:
                if f not in unknown_filters:
                    unknown_filters.append(f)

        if unknown_filters:
            print(f"  [search_url] unrecognised filters (ignored): {unknown_filters}")

        # Build the final Entrez query
        # Publication types use OR (article can be any of the selected types)
        # Other filters (date, species, sex, age) use AND
        parts = [f"({term})"]
        if pub_type_clauses:
            parts.append("(" + " OR ".join(pub_type_clauses) + ")")
        parts.extend(other_clauses)
        entrez_query = " AND ".join(parts)

        # Use clean term as display label (not the full Entrez query with all clauses)
        filter_summary = "; ".join(filters) if filters else "none"
        display_label  = term   # show just the search term, not the full entrez query

        print(f"\n[search_url] ===== DEBUG =====")
        print(f"[search_url] Term: {term}")
        print(f"[search_url] Filters: {filters}")
        print(f"[search_url] Entrez query: {entrez_query}")
        print(f"[search_url] Pub types ({len(pub_type_clauses)}): {pub_type_clauses}")
        print(f"[search_url] Other clauses ({len(other_clauses)}): {other_clauses}")
        print(f"[search_url] Unknown filters: {unknown_filters}")
        print(f"[search_url] ====================")

        # ── Run same pipeline as /search ─────────────────────────────────
        # IMPORTANT: use entrez_query (with all filters) for the actual search
        # but display_label (clean term) for the keyword card display
        search_keywords  = [entrez_query]   # full query with filters → sent to PubMed API
        display_keywords = [display_label]  # clean term → shown in UI cards

        all_results, ncbi_totals = scraper.search_multiple_keywords(
            search_keywords, max_results_per_keyword=max_results
        )
        all_results    = scraper.compute_keyword_scores(all_results, search_keywords)
        sorted_results = scraper.sort_results_by_count(all_results, ascending=True)

        keyword_summary = []
        total_articles  = 0

        for keyword, df in sorted_results.items():
            count = len(df)
            total_articles += count
            records    = df.to_dict('records') if not df.empty else []
            free_count = sum(1 for r in records if r.get('is_free_pmc'))
            ncbi_total = ncbi_totals.get(keyword, count)

            # Use display_label for the card title (not the full Entrez query)
            keyword_summary.append({
                'keyword':    display_label,
                'count':      count,
                'ncbi_total': ncbi_total,
                'fetched':    count,
                'has_more':   ncbi_total > count,
                'free_count': free_count,
                'results':    records,
            })

        search_record = {
            'query':     f"URL search: {term[:60]}{'…' if len(term) > 60 else ''}",
            'count':     total_articles,
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        }
        recent_searches.insert(0, search_record)
        if len(recent_searches) > 50:
            recent_searches.pop()

        total_free_articles = sum(s['free_count'] for s in keyword_summary)
        total_ncbi_articles = sum(s['ncbi_total'] for s in keyword_summary)
        keywords_str        = display_label   # show clean term in UI, not full Entrez query

        search_id = hashlib.md5(
            (entrez_query + str(max_results)).encode()
        ).hexdigest()[:12]
        _search_cache[search_id] = {
            'sorted_results': {kw: df for kw, df in sorted_results.items()},
            'keywords':       display_keywords,
            'sort_order':     'ascending',
            'max_results':    max_results,
        }
        if len(_search_cache) > CACHE_MAX:
            del _search_cache[next(iter(_search_cache))]

        # ── Map pubt.* filters to sidebar checkbox IDs ────────────────────
        PUBT_TO_CHECKBOX = {
            'pubt.clinicalstudy':             'f-type-clinical-study',
            'pubt.clinicaltrial':             'f-type-clinical',
            'pubt.clinicaltrialphaseii':      'f-type-clinical-2',
            'pubt.clinicaltrialphaseiii':     'f-type-clinical-3',
            'pubt.clinicaltrialphaseiv':      'f-type-clinical-4',
            'pubt.meta-analysis':             'f-type-meta',
            'pubt.observationalstudy':        'f-type-observational',
            'pubt.randomizedcontrolledtrial': 'f-type-rct',
            'pubt.review':                    'f-type-review',
            'pubt.systematicreview':          'f-type-systematic',
            'pubt.controlledclinicaltrial':   'f-type-clinical',
            'pubt.comparativestudy':          'f-type-journal',
            'pubt.multicenterstudy':          'f-type-journal',
            'pubt.conferenceproceedings':     'f-type-journal',
            'datesearch.y_1':                 'date_1',
            'datesearch.y_5':                 'date_5',
            'datesearch.y_10':                'date_10',
            'hum_ani.humans':                 'f-humans',
            'hum_ani.animals':                'f-animals',
            'sex.female':                     'f-female',
            'sex.male':                       'f-male',
            'ages.child':                     'f-child',
            'ages.adult':                     'f-adult',
            'ages.aged':                      'f-aged',
            'ages.infant':                    'f-infant',
            'simsearch2.ffrft':               'f-free-pmc',
        }
        PUBT_TO_MODAL_TYPE = {
            'pubt.clinicaltrialphaseii':      'Clinical Trial, Phase II',
            'pubt.clinicaltrialphaseiii':     'Clinical Trial, Phase III',
            'pubt.clinicaltrialphaseiv':      'Clinical Trial, Phase IV',
            'pubt.comparativestudy':          'Comparative Study',
            'pubt.conferenceproceedings':     'Conference Proceedings',
            'pubt.controlledclinicaltrial':   'Controlled Clinical Trial',
            'pubt.multicenterstudy':          'Multicenter Study',
            'pubt.observationalstudy':        'Observational Study',
            'pubt.randomizedcontrolledtrial': 'Randomized Controlled Trial',
            'pubt.systematicreview':          'Systematic Review',
            'pubt.meta-analysis':             'Meta-Analysis',
            'pubt.review':                    'Review',
            'pubt.clinicalstudy':             'Clinical Study',
            'pubt.clinicaltrial':             'Clinical Trial',
        }
        active_checkboxes  = list({PUBT_TO_CHECKBOX[f]    for f in filters if f in PUBT_TO_CHECKBOX})
        active_modal_types = list({PUBT_TO_MODAL_TYPE[f]  for f in filters if f in PUBT_TO_MODAL_TYPE})
        active_date        = next((f.replace('datesearch.y_','') for f in filters if f.startswith('datesearch.y_')), '')
        # Tell the frontend that filters came from URL — skip initial client-side filtering
        url_filters_applied = True

        return render_template('results_multi.html',
                             keywords=display_keywords,
                             keywords_str=keywords_str,
                             search_id=search_id,
                             keyword_summary=keyword_summary,
                             total_articles=total_articles,
                             total_ncbi_articles=total_ncbi_articles,
                             total_free_articles=total_free_articles,
                             active_checkboxes=active_checkboxes,
                             active_modal_types=active_modal_types,
                             active_date=active_date,
                             entrez_query=entrez_query,
                             url_filters_applied=True)

    except Exception as e:
        import traceback; traceback.print_exc()
        return render_template('index_multi_with_logo.html',
                             error=f"An error occurred: {str(e)}",
                             active_tab='url_search',
                             recent_searches=recent_searches[:10])



@app.route('/api/search/multi', methods=['POST'])
def api_search_multi():
    """API endpoint for multi-keyword searching"""
    try:
        data = request.get_json()
        
        keywords = data.get('keywords', [])
        if isinstance(keywords, str):
            keywords = [k.strip() for k in keywords.split(',') if k.strip()]
        
        max_results = int(data.get('max_results_per_keyword', 20))
        sort_order = data.get('sort_order', 'ascending')
        
        if not keywords:
            return jsonify({'error': 'Keywords parameter is required'}), 400
        
        if max_results < 1 or max_results > 10000:
            return jsonify({'error': 'max_results must be between 1 and 10,000'}), 400
        
        # Search all keywords
        all_results, _ = scraper.search_multiple_keywords(keywords, max_results)
        
        # Sort by article count
        sorted_results = scraper.sort_results_by_count(all_results, ascending=(sort_order == 'ascending'))
        
        # Convert to JSON format
        response_data = {
            'total_keywords': len(keywords),
            'sort_order': sort_order,
            'results': []
        }
        
        for keyword, df in sorted_results.items():
            response_data['results'].append({
                'keyword': keyword,
                'count': len(df),
                'articles': df.to_dict('records') if not df.empty else []
            })
        
        return jsonify(response_data)
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/load_more', methods=['POST'])
@login_required
def load_more():
    """
    Fetch the next batch of articles for a single keyword starting at an offset.
    Called by the results page 'Load More' button via AJAX.

    Request JSON:
        { "keyword": "lung cancer", "offset": 500, "batch": 500, "all_keywords": ["lung cancer", "breast cancer"] }

    Response JSON:
        { "keyword": "...", "articles": [...], "fetched": N, "offset": N, "has_more": bool, "ncbi_total": N }
    """
    try:
        data       = request.get_json()
        keyword    = data.get('keyword', '').strip()
        offset     = int(data.get('offset', 500))
        batch      = int(data.get('batch', 500))
        all_kws    = data.get('all_keywords', [keyword])

        if not keyword:
            return jsonify({'error': 'keyword is required'}), 400

        batch = max(1, min(batch, 500))   # cap at 500 per call

        # Fetch next batch
        articles = scraper.fetch_more(keyword, offset=offset, batch=batch)
        articles = [a for a in articles if a is not None]

        # Score the new batch against all keywords
        if articles:
            df_new = pd.DataFrame(articles)
            df_new['keyword'] = keyword

            # Lightweight scoring: keyword_match_count = 1 (single batch),
            # total hits computed via combined regex
            all_terms = list({
                term
                for kw in all_kws
                for term in scraper._terms_from_keyword(kw)
            })
            if all_terms:
                pattern = re.compile(
                    "|".join(re.escape(t) for t in all_terms), re.IGNORECASE
                )
                text_series = (
                    df_new['title'].fillna('') + ' ' + df_new['abstract'].fillna('')
                ).str.lower()
                df_new['keyword_total_hits']  = text_series.apply(
                    lambda t: len(pattern.findall(t))
                )
            else:
                df_new['keyword_total_hits'] = 0

            df_new['keyword_match_count'] = 1   # will update client-side if cross-match

            # Get real NCBI total to know if even more exist
            _, ncbi_total = scraper.search_pubmed(keyword, max_results=1)
            new_offset = offset + len(articles)

            return jsonify({
                'keyword':    keyword,
                'articles':   df_new.to_dict('records'),
                'fetched':    len(articles),
                'offset':     new_offset,
                'ncbi_total': ncbi_total,
                'has_more':   new_offset < ncbi_total,
            })
        else:
            _, ncbi_total = scraper.search_pubmed(keyword, max_results=1)
            return jsonify({
                'keyword':    keyword,
                'articles':   [],
                'fetched':    0,
                'offset':     offset,
                'ncbi_total': ncbi_total,
                'has_more':   False,
            })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


def _presplit_fulltext(df: 'pd.DataFrame') -> 'pd.DataFrame':
    """
    Before writing to Excel, split any 'Full Text (PMC)' values that exceed
    EXCEL_CELL_LIMIT into separate columns: 'Full Text (PMC) - Part 2', etc.

    This must run BEFORE df.to_excel() because openpyxl silently truncates
    any cell value longer than 32,767 chars -- meaning _split_fulltext_across_cells
    would never see the overflowing text if we waited until after writing.
    """
    LIMIT = EXCEL_CELL_LIMIT
    FT_COL = 'Full Text (PMC)'

    if FT_COL not in df.columns:
        return df

    df = df.copy()

    # Find how many parts we need across all rows
    max_parts = 1
    for val in df[FT_COL]:
        if val and isinstance(val, str) and len(val) > LIMIT:
            parts = _count_parts(val, LIMIT)
            max_parts = max(max_parts, parts)

    if max_parts <= 1:
        return df   # nothing to do

    # Add Part columns (initially empty)
    ft_col_pos = df.columns.get_loc(FT_COL)
    for part in range(2, max_parts + 1):
        col_name = f'{FT_COL} - Part {part}'
        if col_name not in df.columns:
            # Insert immediately after the previous part
            insert_pos = ft_col_pos + (part - 1)
            df.insert(insert_pos, col_name, '')

    # Fill each row's parts
    def split_row(val):
        if not val or not isinstance(val, str) or len(val) <= LIMIT:
            return [val] + [''] * (max_parts - 1)
        chunks = _chunk_text(val, LIMIT)
        # Pad to max_parts
        return chunks + [''] * (max_parts - len(chunks))

    parts_data = df[FT_COL].apply(split_row)

    df[FT_COL] = parts_data.apply(lambda x: x[0])
    for i in range(1, max_parts):
        col_name = f'{FT_COL} - Part {i + 1}'
        df[col_name] = parts_data.apply(lambda x, i=i: x[i] if i < len(x) else '')

    return df


def _count_parts(text: str, limit: int) -> int:
    """Count how many chunks a text splits into at the given limit."""
    count = 0
    while text:
        count += 1
        if len(text) <= limit:
            break
        cut = _find_cut(text, limit)
        text = text[cut:].lstrip()
    return count


def _chunk_text(text: str, limit: int) -> list:
    """Split text into chunks of at most `limit` chars, breaking at word boundaries."""
    chunks = []
    while text:
        if len(text) <= limit:
            chunks.append(text)
            break
        cut = _find_cut(text, limit)
        chunks.append(text[:cut].rstrip())
        text = text[cut:].lstrip()
    return chunks


def _find_cut(text: str, limit: int) -> int:
    """Find the best cut point at or before `limit` chars."""
    # Prefer paragraph boundary
    cut = text[:limit].rfind('\n')
    if cut < limit * 0.75:
        # Fall back to word boundary
        cut = text[:limit].rfind(' ')
    if cut < limit * 0.75:
        # Hard cut
        cut = limit
    return cut


def _split_fulltext_across_cells(ws, wrap_align):
    """
    Find 'Full Text (PMC)' column in ws.
    - Text that fits within EXCEL_CELL_LIMIT stays in place.
    - Text that exceeds the limit is split across continuation columns
      ('Full Text (PMC) - Part 2', 'Part 3', ...) added to the right.
    - All parts get wrap_text so the text displays fully in the cell.
    """
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    LIMIT = EXCEL_CELL_LIMIT   # 32,000 chars

    # ── Find Full Text column ────────────────────────────────────────────
    ft_col_idx = None
    for cell in ws[1]:
        if cell.value == 'Full Text (PMC)':
            ft_col_idx = cell.column
            break
    if ft_col_idx is None:
        return

    # ── Style definitions ────────────────────────────────────────────────
    hdr_fill = PatternFill(start_color='1A365D', end_color='1A365D', fill_type='solid')
    hdr_font = Font(color='FFFFFF', bold=True, size=10)
    cell_align = Alignment(wrap_text=True, vertical='top')

    # ── First pass: split every long value into chunks ───────────────────
    row_chunks = {}   # {row_number: [chunk1, chunk2, ...]}
    max_parts  = 1

    for row in ws.iter_rows(min_row=2):
        cell  = row[ft_col_idx - 1]
        value = cell.value

        if not value or not isinstance(value, str):
            # Ensure empty cells still get wrap alignment
            cell.alignment = cell_align
            continue

        if len(value) <= LIMIT:
            # Fits -- just set alignment and move on
            cell.alignment = cell_align
            continue

        # Split into chunks at word boundaries
        chunks   = []
        text     = value
        while text:
            if len(text) <= LIMIT:
                chunks.append(text)
                break
            # Find last space or newline within the limit to avoid cutting mid-word
            cut = text[:LIMIT].rfind('\n')
            if cut < LIMIT * 0.75:
                cut = text[:LIMIT].rfind(' ')
            if cut < LIMIT * 0.75:
                cut = LIMIT    # no good break point -- hard cut
            chunks.append(text[:cut].rstrip())
            text = text[cut:].lstrip()

        row_chunks[cell.row] = chunks
        max_parts = max(max_parts, len(chunks))

    if max_parts <= 1:
        return   # nothing to split

    # ── Add continuation header columns to the right of existing columns ─
    base_col   = ws.max_column   # rightmost col BEFORE we add parts
    part_col_map = {}            # {part_number: col_index}

    for part in range(2, max_parts + 1):
        col_idx = base_col + (part - 1)
        part_col_map[part] = col_idx

        hdr = ws.cell(row=1, column=col_idx)
        hdr.value     = f'Full Text (PMC) - Part {part}'
        hdr.fill      = hdr_fill
        hdr.font      = hdr_font
        hdr.alignment = cell_align
        ws.column_dimensions[get_column_letter(col_idx)].width = 80

    # Also set Part 1 column width
    ws.column_dimensions[get_column_letter(ft_col_idx)].width = 80

    # ── Second pass: write chunks into cells ─────────────────────────────
    for row_num, chunks in row_chunks.items():

        # Detect row background colour (green for free PMC rows)
        orig_fill = None
        orig_cell_fill = ws.cell(row=row_num, column=ft_col_idx).fill
        if (orig_cell_fill
                and orig_cell_fill.fill_type not in (None, 'none')
                and orig_cell_fill.fgColor
                and orig_cell_fill.fgColor.rgb not in ('00000000', 'FF000000', '00FFFFFF', 'FFFFFFFF')):
            orig_fill = orig_cell_fill

        # Part 1 -- overwrite original cell
        c1 = ws.cell(row=row_num, column=ft_col_idx)
        c1.value     = chunks[0]
        c1.alignment = cell_align

        # Parts 2+ -- write into continuation columns
        for part, chunk in enumerate(chunks[1:], start=2):
            col_idx = part_col_map[part]
            c = ws.cell(row=row_num, column=col_idx)
            c.value     = chunk
            c.alignment = cell_align
            if orig_fill:
                c.fill = orig_fill


@app.route('/export/multi/<format>')
@login_required
def export_multi(format):
    """Export results -- reads from cache if available, else re-searches.

    mode (query param):
        'abstract'  → abstract-only columns, no full text fetched (fast)
        'pmc_only'  → free PMC articles only, fetch full text + PDF hits
        'full'      → all columns including full text fetched from PMC (default, slower)

    shown (query param, JSON):
        Per-keyword dict of how many articles are shown on screen after filters.
        If provided, export is limited to exactly those rows per keyword.
        e.g. {"lung cancer": 259, "breast cancer": 254}
    """
    try:
        search_id    = request.args.get('search_id', '')
        keywords_str = request.args.get('keywords', '')
        max_results  = int(request.args.get('max_results', DEFAULT_MAX_RESULTS))
        sort_order   = request.args.get('sort_order', 'ascending')
        mode         = request.args.get('mode', 'full')

        # ── Try cache first ──────────────────────────────────────────────
        if search_id and search_id in _search_cache:
            cached         = _search_cache[search_id]
            sorted_results = cached['sorted_results']
            keywords       = cached['keywords']
            print(f"  [export] cached results  search_id={search_id}  mode={mode}")
        else:
            print(f"  [export] cache miss -- re-searching...  mode={mode}")
            if not keywords_str:
                return "No keywords provided", 400
            keywords = [k.strip() for k in keywords_str.split(',') if k.strip()]
            all_results, _ = scraper.search_multiple_keywords(keywords, max_results)
            all_results    = scraper.compute_keyword_scores(all_results, keywords)
            sorted_results = scraper.sort_results_by_count(
                all_results, ascending=(sort_order == 'ascending')
            )

        # ── Read sidebar filter params from request ──────────────────────
        f_date              = request.args.get('f_date', '').strip()
        f_date_from         = request.args.get('f_date_from', '').strip()
        f_date_to           = request.args.get('f_date_to', '').strip()
        f_free_pmc          = request.args.get('f_free_pmc', '') == '1'
        f_abstract          = request.args.get('f_abstract', '') == '1'
        f_humans            = request.args.get('f_humans', '') == '1'
        f_animals           = request.args.get('f_animals', '') == '1'
        f_female            = request.args.get('f_female', '') == '1'
        f_male              = request.args.get('f_male', '') == '1'
        f_child             = request.args.get('f_child', '') == '1'
        f_adult             = request.args.get('f_adult', '') == '1'
        f_aged              = request.args.get('f_aged', '') == '1'
        f_infant            = request.args.get('f_infant', '') == '1'
        f_type_journal      = request.args.get('f_type_journal', '') == '1'
        f_type_review       = request.args.get('f_type_review', '') == '1'
        f_type_systematic   = request.args.get('f_type_systematic', '') == '1'
        f_type_meta         = request.args.get('f_type_meta', '') == '1'
        f_type_rct          = request.args.get('f_type_rct', '') == '1'
        f_type_clinical     = request.args.get('f_type_clinical', '') == '1'
        f_type_case         = request.args.get('f_type_case', '') == '1'
        f_type_observational = request.args.get('f_type_observational', '') == '1'

        any_filter_active = any([
            f_date, f_free_pmc, f_abstract,
            f_humans, f_animals, f_female, f_male,
            f_child, f_adult, f_aged, f_infant,
            f_type_journal, f_type_review, f_type_systematic,
            f_type_meta, f_type_rct, f_type_clinical,
            f_type_case, f_type_observational,
        ])

        def apply_sidebar_filters(df):
            """Apply the same filters the user selected in the left sidebar."""
            if df.empty or not any_filter_active:
                return df

            mask = pd.Series([True] * len(df), index=df.index)

            # ── Publication date ──────────────────────────────────────────
            def parse_year(d):
                """Extract 4-digit year from date like 'Jan 2025' or '2025'.
                Returns 9999 for unparseable dates so they are always kept."""
                m = re.search(r'\b(19|20)\d{2}\b', str(d))
                return int(m.group()) if m else 9999

            if f_date and f_date != 'custom':
                cutoff = datetime.now().year - int(f_date)
                mask &= df['publication_date'].apply(parse_year) >= cutoff

            elif f_date == 'custom':
                years = df['publication_date'].apply(parse_year)
                if f_date_from:
                    mask &= (years >= int(f_date_from)) | (years == 9999)
                if f_date_to:
                    mask &= (years <= int(f_date_to)) | (years == 9999)

            # ── Text availability ─────────────────────────────────────────
            if f_free_pmc and 'is_free_pmc' in df.columns:
                mask &= df['is_free_pmc'] == True
            if f_abstract:
                mask &= df['abstract'].notna() & (df['abstract'] != 'N/A') & (df['abstract'] != '')

            # ── Article type -- OR logic (keep if matches ANY checked) ──────
            any_type = any([f_type_journal, f_type_review, f_type_systematic,
                            f_type_meta, f_type_rct, f_type_clinical,
                            f_type_case, f_type_observational])
            if any_type:
                pt = df['publication_type'].fillna('').str.lower()
                type_mask = pd.Series([False] * len(df), index=df.index)
                if f_type_journal:      type_mask |= pt.str.contains('journal article', na=False)
                if f_type_review:       type_mask |= pt.str.contains('review', na=False)
                if f_type_systematic:   type_mask |= pt.str.contains('systematic', na=False)
                if f_type_meta:         type_mask |= pt.str.contains('meta-analysis', na=False)
                if f_type_rct:          type_mask |= pt.str.contains('randomized', na=False)
                if f_type_clinical:     type_mask |= pt.str.contains('clinical trial', na=False)
                if f_type_case:         type_mask |= pt.str.contains('case report', na=False)
                if f_type_observational: type_mask |= pt.str.contains('observational', na=False)
                mask &= type_mask

            # ── Species / sex / age -- text-based matching ─────────────────
            text = (df['abstract'].fillna('') + ' ' + df['affiliation'].fillna('')).str.lower()
            if f_humans  and not f_animals: mask &= text.str.contains('human|patient|cohort', na=False, regex=True)
            if f_animals and not f_humans:  mask &= text.str.contains('animal|mouse|rat|murine', na=False, regex=True)
            if f_female  and not f_male:    mask &= text.str.contains('female|women|woman', na=False, regex=True)
            if f_male    and not f_female:  mask &= text.str.contains(r'\bmale\b|\bmen\b', na=False, regex=True)
            if f_child:  mask &= text.str.contains('child|pediatric|adolescent', na=False, regex=True)
            if f_adult:  mask &= text.str.contains('adult', na=False, regex=True)
            if f_aged:   mask &= text.str.contains('elderly|aged|geriatric', na=False, regex=True)
            if f_infant: mask &= text.str.contains('infant|newborn|neonatal', na=False, regex=True)

            return df[mask]

        # ── Apply row limits and sidebar filters ─────────────────────────
        limited_results = {}
        for kw, df in sorted_results.items():
            if df.empty:
                limited_results[kw] = df
                continue
            df = df.head(max_results)
            # pmc_only mode: keep only free PMC rows
            if mode == 'pmc_only' and 'is_free_pmc' in df.columns:
                df = df[df['is_free_pmc'] == True]
            # Apply sidebar filters -- this produces exactly the same
            # article set that is shown on screen
            df = apply_sidebar_filters(df)
            limited_results[kw] = df

        # ── Helper: prepare one DataFrame for export ─────────────────────
        def prepare_df(df):
            """
            Apply column filtering based on mode:
              abstract  → PMID + abstract columns only, NO full text fetched
              pmc_only  → free PMC articles only, fetch full text + PDF hits
              full      → all columns, fetch full text + PDF hits for free PMC
            """
            df = df.copy()

            if mode == 'abstract':
                # ── Only abstract columns -- fast, no network calls ────────
                drop_internal = [c for c in ['pub_type_label', 'pdf_url', 'pmc_url', 'is_free_pmc'] if c in df.columns]
                if drop_internal:
                    df = df.drop(columns=drop_internal)
                ordered = [c for c in ABSTRACT_ONLY_COLUMNS if c in df.columns]
                df = df[ordered]
                rename_map = {
                    'pmid':                'PMID',
                    'title':               'Title',
                    'publication_type':    'Publication Type',
                    'authors':             'Authors',
                    'affiliation':         'Affiliation',
                    'country':             'Country',
                    'journal':             'Journal',
                    'publication_date':    'Publication Date',
                    'abstract':            'Abstract',
                    'pmc_id':              'PMC ID',
                    'url':                 'PubMed URL',
                    'keyword':             'Keyword',
                    'keyword_match_count': 'Keyword Match Count',
                    'matched_keywords':    'Matched Keywords',
                    'per_keyword_hits':    'Hits Per Keyword',
                    'keyword_total_hits':  'Total Keyword Hits',
                }
                df = df.rename(columns={k: v for k, v in rename_map.items() if k in df.columns})

            elif mode == 'pmc_only':
                # ── Free PMC only -- fetch full text, include PDF hit scores ─
                # Row filtering already done above (only is_free_pmc rows)
                # Now fetch full text + PDF keyword scores
                df = add_full_text_column(df, keywords=keywords)
                # Keep PMID, title, abstract, full text, PDF scores
                PMC_COLUMNS = [
                    'keyword', 'pmid', 'title', 'publication_type', 'authors',
                    'journal', 'publication_date', 'abstract',
                    'pmc_id', 'url',
                    'keyword_total_hits', 'per_keyword_hits',
                    'pdf_total_keyword_hits', 'pdf_per_keyword_hits',
                    'Free article complete content',   # LAST
                ]
                drop_internal = [c for c in ['pub_type_label', 'pdf_url', 'pmc_url', 'is_free_pmc'] if c in df.columns]
                if drop_internal:
                    df = df.drop(columns=drop_internal)
                ordered = [c for c in PMC_COLUMNS if c in df.columns]
                df = df[ordered]
                rename_map = {
                    'pmid':                          'PMID',
                    'title':                         'Title',
                    'publication_type':              'Publication Type',
                    'authors':                       'Authors',
                    'journal':                       'Journal',
                    'publication_date':              'Publication Date',
                    'abstract':                      'Abstract',
                    'pmc_id':                        'PMC ID',
                    'url':                           'PubMed URL',
                    'keyword':                       'Keyword',
                    'keyword_total_hits':             'Total Keyword Hits',
                    'per_keyword_hits':              'Hits Per Keyword (Abstract)',
                    'Free article complete content': 'Full Text (PMC)',
                    'pdf_total_keyword_hits':        'PDF Total Keyword Hits',
                    'pdf_per_keyword_hits':          'PDF Hits Per Keyword',
                }
                df = df.rename(columns={k: v for k, v in rename_map.items() if k in df.columns})

            else:
                # ── Full export -- all columns + full text + PDF hits ────────
                df = add_full_text_column(df, keywords=keywords)
                df = clean_for_export(df)

            return df

        # ── Mode label for filename ──────────────────────────────────────
        mode_suffix = {'abstract': '_abstract', 'pmc_only': '_free_pmc', 'full': ''}.get(mode, '')
        timestamp   = datetime.now().strftime('%Y%m%d_%H%M%S')

        # ── CSV ──────────────────────────────────────────────────────────
        if format == 'csv':
            frames = []
            for kw, df in limited_results.items():
                if df.empty:
                    continue
                frames.append(prepare_df(df))

            if not frames:
                return "No results to export", 404

            combined = pd.concat(frames, ignore_index=True)

            # ── Split Full Text into Part 2, Part 3 columns (same as Excel) ──
            # Must happen BEFORE text cleaning so the split logic sees the
            # full original text, not the newline-stripped version.
            combined = _presplit_fulltext(combined)

            # ── Clean ALL text fields ─────────────────────────────────────
            # Replace newlines with space so each article stays on one CSV row.
            # Replace stray quotes with single quote so QUOTE_ALL wraps cleanly.
            for col in combined.columns:
                if combined[col].dtype == object:
                    combined[col] = (
                        combined[col]
                        .fillna('')
                        .astype(str)
                        .str.replace('\r\n', ' ', regex=False)
                        .str.replace('\r',   ' ', regex=False)
                        .str.replace('\n',   ' ', regex=False)
                        .str.replace('"',    "'", regex=False)
                        .str.replace(r'  +', ' ', regex=True)
                        .str.strip()
                    )

            output = io.StringIO()
            combined.to_csv(
                output,
                index=False,
                sep=',',
                quoting=__import__('csv').QUOTE_ALL,
            )
            # utf-8-sig BOM ensures Excel opens with correct encoding and columns
            return send_file(
                io.BytesIO(output.getvalue().encode('utf-8-sig')),
                mimetype='text/csv; charset=utf-8-sig',
                as_attachment=True,
                download_name=f'pubmed{mode_suffix}_{timestamp}.csv'
            )

        # ── JSON ─────────────────────────────────────────────────────────
        elif format == 'json':
            structured = {
                'search_date':    timestamp,
                'export_mode':    mode,
                'sort_order':     sort_order,
                'total_keywords': len(limited_results),
                'keywords':       []
            }
            for kw, df in limited_results.items():
                df_clean = prepare_df(df) if not df.empty else df
                structured['keywords'].append({
                    'keyword':       kw,
                    'article_count': len(df_clean),
                    'articles':      df_clean.to_dict('records') if not df_clean.empty else []
                })

            return send_file(
                io.BytesIO(json.dumps(structured, indent=2).encode()),
                mimetype='application/json',
                as_attachment=True,
                download_name=f'pubmed{mode_suffix}_{timestamp}.json'
            )

        # ── Excel ────────────────────────────────────────────────────────
        elif format == 'excel':
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter

            output = io.BytesIO()
            print(f"📊 Building Excel workbook  mode={mode}...")

            with pd.ExcelWriter(output, engine='openpyxl') as writer:

                hdr_fill   = PatternFill(start_color='1A365D', end_color='1A365D', fill_type='solid')
                hdr_font   = Font(color='FFFFFF', bold=True)
                green_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
                wrap_align = Alignment(wrap_text=True, vertical='top')

                # Summary sheet
                summary_rows = []
                for kw, df in limited_results.items():
                    free_count = int(df['is_free_pmc'].sum()) if 'is_free_pmc' in df.columns else 0
                    summary_rows.append({
                        'Keyword':           kw,
                        'Export Mode':       mode,
                        'Articles Exported': len(df),
                        'Free PMC Articles': free_count,
                    })
                pd.DataFrame(summary_rows).to_excel(writer, sheet_name='Summary', index=False)

                all_frames = []

                for kw, df in limited_results.items():
                    if df.empty:
                        continue

                    # Free PMC first within each sheet
                    if 'is_free_pmc' in df.columns:
                        df = df.sort_values('is_free_pmc', ascending=False)

                    df_clean = prepare_df(df)
                    all_frames.append(df_clean)

                    # Sheet name (max 31 chars, no special chars)
                    sheet_name = "".join(
                        c for c in kw if c.isalnum() or c in (' ', '-', '_')
                    )[:31]

                    # ── Pre-split Full Text BEFORE writing to avoid silent truncation ──
                    df_clean = _presplit_fulltext(df_clean)
                    df_clean.to_excel(writer, sheet_name=sheet_name, index=False)

                    # ── Apply continuation cell formatting (Part 2, Part 3 cols) ──
                    _split_fulltext_across_cells(writer.sheets[sheet_name], wrap_align)

                    # ── Styling ──────────────────────────────────────────
                    ws = writer.sheets[sheet_name]

                    # Style header row
                    for cell in ws[1]:
                        cell.fill      = hdr_fill
                        cell.font      = hdr_font
                        cell.alignment = wrap_align

                    # Find key column indices from header row
                    col_map      = {cell.value: cell.column for cell in ws[1]}
                    pmc_col      = col_map.get('PMC ID')
                    fulltext_col = col_map.get('Full Text (PMC)')
                    abstract_col = col_map.get('Abstract')

                    for row in ws.iter_rows(min_row=2):
                        is_free = bool(pmc_col and row[pmc_col - 1].value not in (None, 'N/A', ''))
                        if is_free:
                            for cell in row:
                                cell.fill = green_fill
                        # Wrap text for Full Text and Abstract columns
                        if fulltext_col:
                            row[fulltext_col - 1].alignment = wrap_align
                        if abstract_col:
                            row[abstract_col - 1].alignment = wrap_align

                    # Column widths
                    width_map = {
                        'Full Text (PMC)': 100,
                        'Abstract':        60,
                        'Affiliation':     40,
                        'Title':           50,
                        'Publication Type': 30,
                        'Authors':         35,
                        'Journal':         30,
                        'PubMed URL':      35,
                        'PDF Hits Per Keyword': 35,
                        'Hits Per Keyword (Abstract)': 35,
                    }
                    for col_cells in ws.columns:
                        hdr = str(col_cells[0].value or '')
                        # Part columns get same width as Full Text
                        if hdr.startswith('Full Text (PMC) - Part'):
                            w = 100
                        else:
                            w = width_map.get(hdr, 18)
                        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = w

                    # Row heights -- taller for rows that have full text content
                    for row in ws.iter_rows(min_row=2):
                        has_ft = fulltext_col and row[fulltext_col - 1].value
                        # Set a generous row height so wrapped text is visible
                        # Excel will still allow users to resize rows manually
                        ws.row_dimensions[row[0].row].height = 200 if has_ft else 60

                # All Results sheet
                if all_frames:
                    all_combined = pd.concat(all_frames, ignore_index=True)
                    all_combined  = _presplit_fulltext(all_combined)
                    all_combined.to_excel(writer, sheet_name='All Results', index=False)
                    _split_fulltext_across_cells(
                        writer.sheets['All Results'], wrap_align
                    )

            output.seek(0)
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'pubmed{mode_suffix}_{timestamp}.xlsx'
            )

        else:
            return "Invalid format", 400

    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        print(f"[export_multi] ERROR: {e}\n{tb}")
        # Return detailed error so we can diagnose on Render
        return jsonify({
            'error': str(e),
            'traceback': tb[-500:]   # last 500 chars of traceback
        }), 500

@app.route('/about')
def about():
    """About page"""
    return render_template('about.html')


@app.route('/api/health')
def health():
    """Health check endpoint"""
    return jsonify({
        'status': 'healthy',
        'email': NCBI_EMAIL,
        'features': ['multi-keyword-search', 'sorted-results', 'copilot'],
        'timestamp': datetime.now().isoformat()
    })


@app.route('/analyze')
@login_required
def analyze():
    """Analyze & report page -- wraps current search with AI copilot."""
    search_id    = request.args.get('search_id', '')
    keywords_str = request.args.get('keywords', '')
    max_results  = int(request.args.get('max_results', DEFAULT_MAX_RESULTS))
    sort_order   = request.args.get('sort_order', 'ascending')

    if search_id and search_id in _search_cache:
        cached   = _search_cache[search_id]
        keywords = cached['keywords']
        total    = sum(len(df) for df in cached['sorted_results'].values())
    else:
        keywords = [k.strip() for k in keywords_str.split(',') if k.strip()]
        total    = 0

    return render_template(
        'analyze.html',
        search_id    = search_id,
        keywords_str = keywords_str,
        keywords     = keywords,
        max_results  = max_results,
        sort_order   = sort_order,
        total_articles = total,
    )


# ── In-memory extract jobs ──────────────────────────────────────────────
_extract_jobs: dict = {}  # job_id -> {status, done, total, rows, error, filename}
_pdf_jobs:     dict = {}  # job_id -> {status, done, zip_path, error}


def _run_pdf_job(job_id: str, pmc_articles: list, tmp_path: str):
    """Background thread: download/generate PDFs and write ZIP."""
    import zipfile as _zf
    import uuid as _uuid

    job = _pdf_jobs[job_id]
    total = len(pmc_articles)
    job['total'] = total

    browser_hdrs = {
        'User-Agent': f'EpiLite/1.0 ({NCBI_EMAIL})',
        'Accept': 'application/pdf,*/*',
    }

    def _safe_fn(pmid, title):
        clean = re.sub(r'[^\w\s-]', '', title)[:50].strip()
        clean = re.sub(r'\s+', '_', clean)
        return f"PMID_{pmid}_{clean}.pdf"

    def _fetch_pdf(art):
        pmc_id = art.get('pmc_id', '')
        fname  = _safe_fn(art['pmid'], art['title'])
        if pmc_id and pmc_id != 'N/A':
            try:
                _r = http_requests.get(
                    f"https://www.ncbi.nlm.nih.gov/pmc/utils/oa/oa.fcgi?id={pmc_id}",
                    timeout=6, headers=browser_hdrs)
                if _r.status_code == 200:
                    from xml.etree import ElementTree as _ET
                    for _lnk in _ET.fromstring(_r.content).iter('link'):
                        if _lnk.get('format') == 'pdf':
                            _url = _lnk.get('href','').replace('ftp://','https://')
                            _r2  = http_requests.get(_url, timeout=20,
                                                     headers=browser_hdrs, allow_redirects=True)
                            if _r2.status_code == 200 and _r2.content[:4] == b'%PDF':
                                return fname, _r2.content, 'NCBI OA FTP'
                            break
            except Exception:
                pass
        if pmc_id and pmc_id != 'N/A':
            try:
                _r = http_requests.get(
                    f"https://europepmc.org/backend/ptpmcrender.fcgi?accid={pmc_id}&blobtype=pdf",
                    timeout=15, headers=browser_hdrs, allow_redirects=True)
                if _r.status_code == 200 and _r.content[:4] == b'%PDF':
                    return fname, _r.content, 'Europe PMC'
            except Exception:
                pass
        return fname, None, None

    try:
        job['status'] = f'Fetching {total} PDFs in parallel...'
        real_pdfs = {}
        with ThreadPoolExecutor(max_workers=8) as pool:
            futs = {pool.submit(_fetch_pdf, a): a for a in pmc_articles}
            done_count = 0
            for fut in as_completed(futs):
                done_count += 1
                try:
                    fn, pb, src = fut.result()
                    real_pdfs[fn] = (pb, src)
                except Exception:
                    pass
                job['status']  = f'Fetching PDFs... {done_count}/{total}'
                job['current'] = done_count

        # Write ZIP
        job['status'] = 'Building ZIP file...'
        index_lines = [
            f"EpiLite PDF Export -- {total} articles",
            f"Downloaded: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            "=" * 60, ""
        ]

        with _zf.ZipFile(tmp_path, mode='w',
                         compression=_zf.ZIP_DEFLATED,
                         allowZip64=True) as zf:
            for i, art in enumerate(pmc_articles, 1):
                pmc_id = art.get('pmc_id', '')
                fname  = _safe_fn(art['pmid'], art['title'])
                pdf_bytes, source = real_pdfs.get(fname, (None, None))

                if pdf_bytes:
                    zf.writestr(fname, pdf_bytes)
                    index_lines.append(f"[REAL PDF]  {fname}  ({source})")
                else:
                    # Generate from PMC XML
                    try:
                        full_text = fetch_pdf_text(art['pdf_url']) if art.get('pdf_url') else ''
                        from reportlab.lib.pagesizes import A4
                        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                        from reportlab.lib.units import mm
                        from reportlab.lib import colors
                        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, HRFlowable
                        from reportlab.lib.enums import TA_JUSTIFY
                        buf = io.BytesIO()
                        doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=20*mm, rightMargin=20*mm,
                            topMargin=20*mm, bottomMargin=20*mm)
                        st  = getSampleStyleSheet()
                        def _sty(n, **kw): return ParagraphStyle(n, parent=st['Normal'], **kw)
                        t_s = _sty('T', fontSize=14, fontName='Helvetica-Bold',
                                   textColor=colors.HexColor('#1a365d'), spaceAfter=6, leading=18)
                        m_s = _sty('M', fontSize=9, fontName='Helvetica',
                                   textColor=colors.HexColor('#4a5568'), spaceAfter=3, leading=13)
                        s_s = _sty('S', fontSize=10, fontName='Helvetica-Bold',
                                   textColor=colors.HexColor('#2b6cb0'), spaceBefore=8, spaceAfter=4)
                        b_s = _sty('B', fontSize=9.5, fontName='Helvetica',
                                   textColor=colors.HexColor('#2d3748'),
                                   alignment=TA_JUSTIFY, spaceAfter=6, leading=14)
                        def _p(txt, sty):
                            try:
                                t = str(txt or '').replace('&','&amp;').replace('<','&lt;').replace('>','&gt;')
                                return Paragraph(t, sty)
                            except:
                                return Spacer(1, 2*mm)
                        story = [
                            _p('<font color="#2b6cb0">EpiLite</font> -- Generated PDF', m_s),
                            HRFlowable(width='100%', thickness=2,
                                      color=colors.HexColor('#2b6cb0'), spaceAfter=8),
                            _p(art.get('title',''), t_s),
                        ]
                        for lbl, val in [('PMID',art.get('pmid','')),('PMC ID',pmc_id),
                                         ('Journal',art.get('journal','')),
                                         ('Date',art.get('publication_date','')),
                                         ('Authors',art.get('authors','')),
                                         ('URL',art.get('url',''))]:
                            if val and val not in ('N/A',''):
                                story.append(_p(f'<b>{lbl}:</b> {val}', m_s))
                        abstract = art.get('abstract','') or ''
                        if abstract and abstract != 'N/A':
                            story += [Spacer(1,4*mm),
                                      HRFlowable(width='100%',thickness=0.5,
                                                 color=colors.HexColor('#e2e8f0'),spaceAfter=4),
                                      _p('Abstract', s_s), _p(abstract, b_s)]
                        if full_text:
                            story += [Spacer(1,4*mm),
                                      HRFlowable(width='100%',thickness=0.5,
                                                 color=colors.HexColor('#e2e8f0'),spaceAfter=4),
                                      _p('Full Text (PMC)', s_s)]
                            for blk in full_text.split('\n\n'):
                                blk = blk.strip()
                                if not blk: continue
                                for ln in blk.split('\n'):
                                    ln = ln.strip()
                                    if not ln: story.append(Spacer(1,2*mm))
                                    elif len(ln)<80 and ln==ln.upper() and len(ln)>3:
                                        story.append(_p(f'<b>{ln}</b>', s_s))
                                    elif len(ln)>3: story.append(_p(ln, b_s))
                                story.append(Spacer(1,1*mm))
                        doc.build(story)
                        zf.writestr(fname, buf.getvalue())
                        index_lines.append(f"[GENERATED] {fname}")
                    except Exception as _e:
                        note = fname.replace('.pdf','_info.txt')
                        zf.writestr(note, f"Title: {art.get('title','')}\nPMID: {art.get('pmid','')}\nURL: {art.get('url','')}\n")
                        index_lines.append(f"[ERR] {fname} ({_e})")

            zf.writestr('_index.txt', '\n'.join(index_lines))

        job['done']     = True
        job['zip_path'] = tmp_path
        job['status']   = f'Complete -- {total} PDFs ready'

    except Exception as e:
        import traceback; traceback.print_exc()
        job['done']  = True
        job['error'] = str(e)


@app.route('/export/pdfs')
@login_required
def export_pdfs():
    """
    Start a background PDF export job and return job_id immediately.
    Client polls /export/pdfs/progress/<job_id> then downloads from /export/pdfs/download/<job_id>.
    This avoids Render's 30-second request timeout.
    """
    import uuid as _uuid, tempfile as _tmp

    try:
        search_id    = request.args.get('search_id', '')
        keywords_str = request.args.get('keywords', '')
        max_results  = int(request.args.get('max_results', DEFAULT_MAX_RESULTS))
        sort_order   = request.args.get('sort_order', 'ascending')

        # Read sidebar filters
        f_date            = request.args.get('f_date', '').strip()
        f_date_from       = request.args.get('f_date_from', '').strip()
        f_date_to         = request.args.get('f_date_to', '').strip()
        f_abstract        = request.args.get('f_abstract', '') == '1'
        f_type_journal    = request.args.get('f_type_journal', '') == '1'
        f_type_review     = request.args.get('f_type_review', '') == '1'
        f_type_systematic = request.args.get('f_type_systematic', '') == '1'
        f_type_meta       = request.args.get('f_type_meta', '') == '1'
        f_type_rct        = request.args.get('f_type_rct', '') == '1'
        f_type_clinical   = request.args.get('f_type_clinical', '') == '1'
        f_type_case       = request.args.get('f_type_case', '') == '1'
        f_type_observational = request.args.get('f_type_observational', '') == '1'
        f_humans          = request.args.get('f_humans', '') == '1'
        f_animals         = request.args.get('f_animals', '') == '1'
        f_female          = request.args.get('f_female', '') == '1'
        f_male            = request.args.get('f_male', '') == '1'
        f_child           = request.args.get('f_child', '') == '1'
        f_adult           = request.args.get('f_adult', '') == '1'
        f_aged            = request.args.get('f_aged', '') == '1'
        f_infant          = request.args.get('f_infant', '') == '1'

        # Load from cache or re-search
        if search_id and search_id in _search_cache:
            cached         = _search_cache[search_id]
            sorted_results = cached['sorted_results']
        else:
            if not keywords_str:
                return jsonify({'error': 'No keywords provided'}), 400
            keywords = [k.strip() for k in keywords_str.split(',') if k.strip()]
            all_results, _ = scraper.search_multiple_keywords(keywords, max_results)
            all_results    = scraper.compute_keyword_scores(all_results, keywords)
            sorted_results = scraper.sort_results_by_count(
                all_results, ascending=(sort_order == 'ascending'))

        # Collect free PMC articles with sidebar filters applied
        def parse_year(d):
            m = re.search(r'\b(19|20)\d{2}\b', str(d))
            return int(m.group()) if m else 9999

        seen_pmids   = set()
        pmc_articles = []

        for kw, df in sorted_results.items():
            if df.empty:
                continue
            df = df.head(max_results)
            if 'is_free_pmc' in df.columns:
                df = df[df['is_free_pmc'] == True]

            # Apply date filter
            if f_date and f_date != 'custom':
                cutoff = datetime.now().year - int(f_date)
                df = df[df['publication_date'].apply(parse_year) >= cutoff]
            elif f_date == 'custom':
                years = df['publication_date'].apply(parse_year)
                if f_date_from: df = df[years >= int(f_date_from)]
                if f_date_to:   df = df[years <= int(f_date_to)]

            # Apply type filter
            any_type = any([f_type_journal, f_type_review, f_type_systematic,
                            f_type_meta, f_type_rct, f_type_clinical,
                            f_type_case, f_type_observational])
            if any_type:
                pt = df['publication_type'].fillna('').str.lower()
                tm = pd.Series([False]*len(df), index=df.index)
                if f_type_journal:      tm |= pt.str.contains('journal article', na=False)
                if f_type_review:       tm |= pt.str.contains('review', na=False)
                if f_type_systematic:   tm |= pt.str.contains('systematic', na=False)
                if f_type_meta:         tm |= pt.str.contains('meta-analysis', na=False)
                if f_type_rct:          tm |= pt.str.contains('randomized', na=False)
                if f_type_clinical:     tm |= pt.str.contains('clinical trial', na=False)
                if f_type_case:         tm |= pt.str.contains('case report', na=False)
                if f_type_observational: tm |= pt.str.contains('observational', na=False)
                df = df[tm]

            for _, row in df.iterrows():
                pmid = str(row.get('pmid', ''))
                if pmid in seen_pmids:
                    continue
                seen_pmids.add(pmid)
                pmc_id = str(row.get('pmc_id', '') or '')
                pdf_url = f"https://www.ncbi.nlm.nih.gov/pmc/articles/{pmc_id}/pdf/" if pmc_id else ''
                pmc_articles.append({
                    'pmid':             pmid,
                    'title':            str(row.get('title', '') or ''),
                    'pmc_id':           pmc_id,
                    'pdf_url':          pdf_url,
                    'abstract':         str(row.get('abstract', '') or ''),
                    'authors':          str(row.get('authors', '') or ''),
                    'journal':          str(row.get('journal', '') or ''),
                    'publication_date': str(row.get('publication_date', '') or ''),
                    'url':              str(row.get('url', '') or ''),
                })

        if not pmc_articles:
            return jsonify({'error': 'No free PMC articles found for current filters'}), 404

        # Create temp file for ZIP
        tmp_fd, tmp_path = tempfile.mkstemp(suffix='.zip')
        os.close(tmp_fd)

        # Start background job
        job_id = _uuid.uuid4().hex[:12]
        _pdf_jobs[job_id] = {
            'status':   'Starting...',
            'done':     False,
            'current':  0,
            'total':    len(pmc_articles),
            'zip_path': None,
            'zip_name': f'pubmed_pdfs_{len(pmc_articles)}articles_{datetime.now().strftime("%Y%m%d_%H%M%S")}.zip',
            'error':    None,
        }

        t = threading.Thread(target=_run_pdf_job,
                             args=(job_id, pmc_articles, tmp_path), daemon=True)
        t.start()

        return jsonify({'job_id': job_id, 'total': len(pmc_articles)})

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/export/pdfs/progress/<job_id>')
@login_required
def pdf_progress(job_id):
    job = _pdf_jobs.get(job_id)
    if not job:
        return jsonify({'error': 'Job not found'}), 404
    return jsonify({
        'status':  job['status'],
        'done':    job['done'],
        'current': job['current'],
        'total':   job['total'],
        'error':   job['error'],
    })


@app.route('/export/pdfs/download/<job_id>')
@login_required
def pdf_download(job_id):
    job = _pdf_jobs.get(job_id)
    if not job or not job['done'] or not job['zip_path']:
        return "Job not ready", 404
    if job['error']:
        return f"Job failed: {job['error']}", 500
    return send_file(
        job['zip_path'],
        mimetype='application/zip',
        as_attachment=True,
        download_name=job['zip_name'],
    )

# ── Exact column order for the output Excel ────────────────────────────
EXTRACT_COLUMNS = [
    "Author, year", "Country", "Study title", "Published year",
    "Study type", "Trial phase", "Tumor type", "Cancer name",
    "Disease definition", "Stage/SEER Stage", "",
    "Sample size", "Mean or median age (Yr)", "",
    "Race/ethnicity", "Mon/Combo", "Drug Class", "Treatment",
    "Dosage/strength", "Target", "",
    "Median Follow-up period", "",
    "Duration of treatment (M)", "Time to AE onset", "SAE type",
    "Prophylactics considered during treatment",
    "AE reporting method (NCI CTCAE version 5.0 or NCI CTCAE version 4.0 or NCI CTCAE version 4.1)",
    "AE reporting criteria", "Event by Organ Class", "AE reported Name",
    "Grade I (Mild) - %", "Grade I (Mild) - Denominator",
    "Grade II (Moderate) - %", "Grade II (Moderate) - Denominator",
    "Grade I-II (Mild Moderate+) - %", "Grade I-II (Mild -Moderate+) - Denominator",
    "Grade III (Severe) -%", "Grade III (Severe) - Denominator",
    "Grade III-IV (Severe) -%", "Grade III -IV(Severe) - Denominator",
    "Grade III+  (Severe) -%", "Grade III+  (Severe) -Denominator",
    "Grade IV (Life threatening) - %", "Grade IV (Life threatening) - Denominator",
    "Grade V (Death) - %", "Grade V (Death) - Denominator",
    "All Grade - %", "All Grade -Denominator",
    "Bibliography", "URL",
]

# JSON field keys matching EXTRACT_COLUMNS (blanks stay blank)
EXTRACT_KEYS = [
    "author_year", "country", "study_title", "published_year",
    "study_type", "trial_phase", "tumor_type", "cancer_name",
    "disease_definition", "stage_seer", "",
    "sample_size", "mean_median_age", "",
    "race_ethnicity", "mon_combo", "drug_class", "treatment",
    "dosage_strength", "target", "",
    "median_followup", "",
    "duration_treatment_m", "time_to_ae_onset", "sae_type",
    "prophylactics", "ae_reporting_method",
    "ae_reporting_criteria", "event_organ_class", "ae_name",
    "grade1_pct", "grade1_denom",
    "grade2_pct", "grade2_denom",
    "grade12_pct", "grade12_denom",
    "grade3_pct", "grade3_denom",
    "grade34_pct", "grade34_denom",
    "grade3plus_pct", "grade3plus_denom",
    "grade4_pct", "grade4_denom",
    "grade5_pct", "grade5_denom",
    "all_grade_pct", "all_grade_denom",
    "bibliography", "url",
]


def _make_fallback(title, authors, pub_date, country, journal, url, pmid):
    """Create a metadata-only row when AI extraction fails."""
    year = ''
    m = re.search(r'\b(19|20)\d{2}\b', str(pub_date))
    if m: year = m.group()
    # Format author_year: "LastName et al., YYYY"
    first_author = authors.split(',')[0].split()[-1] if authors else ''
    author_year  = f"{first_author} et al., {year}" if first_author and year else authors[:40]
    return {
        'author_year':    author_year,
        'country':        country,
        'study_title':    title,
        'published_year': year,
        'journal':        journal,
        'url':            url,
        'bibliography':   f"{authors[:60]}. {title[:80]}. {journal}. {year}.",
    }


def _run_extract_job(job_id: str, articles: list):
    """Background thread: extract clinical fields from each article via Claude."""
    job = _extract_jobs[job_id]
    job['total'] = len(articles)
    all_rows = []

    # Azure OpenAI configured via environment variables
    # ── Check API key upfront ─────────────────────────────────────────────
    if not os.environ.get('AZURE_OPENAI_KEY', ''):
        job['error'] = 'AZURE_OPENAI_KEY not set. Please add it to Render environment variables.'
        job['done']  = True
        return

    print(f"[extract] Starting job {job_id} -- {len(articles)} articles")

    for i, art in enumerate(articles, 1):
        if job.get('cancelled'):
            break
        job['current'] = i

        # ── Normalise column names -- exact names from EpiLite export ────
        def _get(d, *keys):
            for k in keys:
                v = d.get(k, '') or d.get(k.lower(), '') or ''
                if v and str(v).strip() not in ('', 'nan', 'None'):
                    return str(v).strip()
            return ''

        title    = _get(art, 'Title', 'title', 'Study title')
        abstract = _get(art, 'Abstract', 'abstract')
        # Use merged full text if available, else individual parts
        fulltext = _get(art, 'full_text_merged',
                             'Full Text (PMC)',
                             'full_text', 'Free article complete content')
        if not fulltext:
            # Manually join parts if merge wasn't done
            parts = [_get(art, f'Full Text (PMC) - Part {n}') for n in range(2, 8)]
            base  = _get(art, 'Full Text (PMC)')
            fulltext = ' '.join(p for p in [base] + parts if p)

        authors  = _get(art, 'Authors', 'authors')
        journal  = _get(art, 'Journal', 'journal')
        pub_date = _get(art, 'Publication Date', 'publication_date')
        pub_type = _get(art, 'Publication Type', 'publication_type')
        country  = _get(art, 'Country', 'country')
        url      = _get(art, 'PubMed URL', 'url', 'URL')
        pmid     = _get(art, 'PMID', 'pmid')

        if not title and not abstract:
            print(f"  [extract] Article {i}: no title or abstract -- skipping. Keys: {list(art.keys())[:8]}")
            continue

        job['status'] = f"Extracting article {i} of {len(articles)}: {title[:60]}…"
        print(f"  [extract] Article {i}: {title[:60]}")

        # For PDFs use full text; for CSV/Excel use abstract + fulltext
        if art.get('_source') == 'pdf':
            content = fulltext[:4000] if fulltext else abstract[:2000]
        else:
            content = (abstract + '\n\n' + fulltext[:3000]) if fulltext else abstract

        # Truncate content smartly -- keep most relevant parts
        max_content = 2500
        if len(content) > max_content:
            # Keep first 1500 (abstract/intro) + last 1000 (results/conclusion)
            content = content[:1500] + '\n...\n' + content[-1000:]

        # Build pre-filled defaults
        year         = re.search(r'\b(19|20)\d{2}\b', pub_date).group() if re.search(r'\b(19|20)\d{2}\b', pub_date) else ''
        first_author = authors.split(',')[0].strip().split()[-1] if authors else ''
        author_year  = (first_author + ' et al., ' + year).strip(' ,') if first_author else authors[:40]
        bibliography = authors[:60] + '. ' + title[:80] + '. ' + journal + '. ' + year + '.'

        # Use clean f-string with JSON template - no string concatenation
        safe_title = title[:80].replace('"', "'").replace('\n', ' ')
        safe_bib   = bibliography.replace('"', "'").replace('\n', ' ')
        safe_url   = url.replace('"', '')

        system_msg = 'You are a clinical research data extractor. Return ONLY valid JSON with a rows array. No markdown, no explanation.'

        user_msg = f"""Extract all clinical fields from this article. Fill every field you can find. Return ONLY JSON.

Title: {title}
Authors: {authors}
Journal: {journal} | Date: {pub_date} | Country: {country} | PMID: {pmid}
URL: {url}

Content:
{content}

Return this JSON (fill in the empty strings with real values from the article):
{{
  "rows": [
    {{
      "author_year": "{author_year}",
      "country": "{country}",
      "study_title": "{safe_title}",
      "published_year": "{year}",
      "study_type": "",
      "trial_phase": "",
      "tumor_type": "",
      "cancer_name": "",
      "disease_definition": "",
      "stage_seer": "",
      "sample_size": "",
      "mean_median_age": "",
      "race_ethnicity": "",
      "mon_combo": "",
      "drug_class": "",
      "treatment": "",
      "dosage_strength": "",
      "target": "",
      "median_followup": "",
      "duration_treatment_m": "",
      "time_to_ae_onset": "",
      "sae_type": "",
      "prophylactics": "",
      "ae_reporting_method": "",
      "ae_reporting_criteria": "",
      "event_organ_class": "",
      "ae_name": "",
      "grade1_pct": "",
      "grade1_denom": "",
      "grade2_pct": "",
      "grade2_denom": "",
      "grade12_pct": "",
      "grade12_denom": "",
      "grade3_pct": "",
      "grade3_denom": "",
      "grade34_pct": "",
      "grade34_denom": "",
      "grade3plus_pct": "",
      "grade3plus_denom": "",
      "grade4_pct": "",
      "grade4_denom": "",
      "grade5_pct": "",
      "grade5_denom": "",
      "all_grade_pct": "",
      "all_grade_denom": "",
      "bibliography": "{safe_bib}",
      "url": "{safe_url}"
    }}
  ]
}}

If multiple adverse events are reported, add more rows with the same article fields but different ae_name and grade fields."""

        try:
            # ── Call Groq with retry on rate limit ────────────────────
            resp = None
            for attempt in range(4):
                resp = http_requests.post(
                    f"{os.environ.get('AZURE_OPENAI_ENDPOINT','').rstrip('/')}/openai/deployments/{os.environ.get('AZURE_OPENAI_DEPLOYMENT','gpt-4o-mini')}/chat/completions?api-version=2024-02-01",
                    headers={
                        'Content-Type':  'application/json',
                        'api-key':       os.environ.get('AZURE_OPENAI_KEY',''),
                    },
                    json={
                        'max_tokens':  2000,
                        'temperature': 0,
                        'messages':    [
                            {'role': 'system', 'content': system_msg},
                            {'role': 'user',   'content': user_msg},
                        ],
                    },
                    timeout=45,
                )
                if resp.status_code in (429, 503):
                    wait = 10 * (attempt + 1)
                    print(f"    Rate limit/unavailable -- waiting {wait}s (attempt {attempt+1}/4)")
                    job['status'] = f"Retrying... ({i}/{len(articles)})"
                    time.sleep(wait)
                else:
                    break

            print(f"    API status: {resp.status_code}")
            if resp.status_code == 200:
                text = resp.json().get('choices', [{}])[0].get('message', {}).get('content', '{}')
                print(f"    Raw response (first 300): {text[:300]}")

                # Strip markdown fences
                text = re.sub(r'^```[a-z]*\n?', '', text.strip())
                text = re.sub(r'\n?```$', '', text.strip())
                # Find JSON object in response
                json_match = re.search(r'\{.*\}', text, re.DOTALL)
                if json_match:
                    text = json_match.group()

                try:
                    extracted = json.loads(text)
                    rows = extracted.get('rows', [])
                    if rows:
                        good_rows = []
                        for row in rows:
                            filled = sum(1 for k, v in row.items()
                                        if v and v not in ('', 'null', 'None')
                                        and k not in ('url',))
                            if filled >= 3:
                                good_rows.append(row)
                        if good_rows:
                            print(f"    Extracted {len(good_rows)} good row(s)")
                            all_rows.extend(good_rows)
                        else:
                            print(f"    Rows too sparse -- metadata fallback")
                            all_rows.append(_make_fallback(title, authors, pub_date, country, journal, url, pmid))
                    else:
                        print(f"    No rows -- metadata fallback")
                        all_rows.append(_make_fallback(title, authors, pub_date, country, journal, url, pmid))
                except json.JSONDecodeError as je:
                    print(f"    JSON parse error: {je} | text: {text[:200]}")
                    all_rows.append(_make_fallback(title, authors, pub_date, country, journal, url, pmid))
            else:
                err = resp.json().get('error', {}).get('message', resp.text[:200]) if resp else 'No response'
                print(f"    API error {resp.status_code if resp else '?'}: {err}")
                all_rows.append(_make_fallback(title, authors, pub_date, country, journal, url, pmid))
        except Exception as e:
            print(f"    Exception: {e}")
            all_rows.append(_make_fallback(title, authors, pub_date, country, journal, url, pmid))

        time.sleep(1.5)  # Azure rate limiting buffer

    # ── Build Excel ───────────────────────────────────────────────────────
    print(f"[extract] Building Excel: {len(all_rows)} rows collected")
    for i2, r in enumerate(all_rows[:3]):
        print(f"  Row {i2+1} keys: {list(r.keys())[:8]}")
        print(f"  Row {i2+1} sample: {dict(list(r.items())[:4])}")

    if not all_rows:
        job['error'] = 'No data extracted from any article. Check terminal for API errors.'
        job['done']  = True
        return

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb  = Workbook()
        ws  = wb.active
        ws.title = 'AE Extraction'

        hdr_fill  = PatternFill('solid', start_color='1A365D')
        hdr_font  = Font(name='Arial', bold=True, color='FFFFFF', size=9)
        hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin = Border(
            left=Side(style='thin', color='FFFFFF'),
            right=Side(style='thin', color='FFFFFF'),
        )

        col_idx = 1
        for col_name in EXTRACT_COLUMNS:
            cell = ws.cell(row=1, column=col_idx, value=col_name if col_name else '')
            if col_name:
                cell.fill  = hdr_fill
                cell.font  = hdr_font
                cell.alignment = hdr_align
                cell.border = thin
            else:
                cell.fill = PatternFill('solid', start_color='2B4C7E')
            col_idx += 1

        ws.freeze_panes = 'A2'

        for r_idx, row_data in enumerate(all_rows, start=2):
            col_idx = 1
            for col_name, key in zip(EXTRACT_COLUMNS, EXTRACT_KEYS):
                val = row_data.get(key, '') if key else ''
                cell = ws.cell(row=r_idx, column=col_idx, value=str(val) if val else '')
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                cell.font = Font(name='Arial', size=9)
                if r_idx % 2 == 0:
                    cell.fill = PatternFill('solid', start_color='EBF4FF')
                col_idx += 1

        width_map = {
            1: 20, 2: 12, 3: 35, 4: 10, 5: 15, 6: 12, 7: 15, 8: 18,
            9: 25, 10: 15, 12: 10, 13: 15, 15: 18, 16: 12, 17: 18,
            18: 20, 19: 18, 20: 15, 22: 18, 24: 18, 25: 15, 26: 20,
            27: 30, 28: 45, 29: 20, 30: 22, 31: 25,
        }
        for c2 in range(1, len(EXTRACT_COLUMNS) + 1):
            ws.column_dimensions[get_column_letter(c2)].width = width_map.get(c2, 12)

        ws.row_dimensions[1].height = 60

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        job['excel_bytes'] = buf.getvalue()
        job['done']   = True
        job['status'] = f'Complete -- {len(all_rows)} rows extracted from {len(articles)} articles'

    except Exception as e:
        job['error'] = f'Excel build failed: {e}'
        job['done']  = True
        import traceback; traceback.print_exc()


@app.route('/api/copilot_file', methods=['POST'])
@login_required
def copilot_file():
    """Copilot chat using uploaded file as context."""
    try:
        f        = request.files.get('file')
        history  = json.loads(request.form.get('history', '[]'))
        question = request.form.get('question', '')
        if not f or not question:
            return jsonify({'error': 'File and question required'}), 400
        fname = f.filename or ''
        articles = []
        try:
            if fname.endswith('.csv'):
                df = pd.read_csv(f, encoding='utf-8-sig', dtype=str).fillna('')
                articles = df.to_dict(orient='records')
            elif fname.endswith(('.xlsx', '.xls')):
                xl = pd.ExcelFile(f)
                sheet = 'All Results' if 'All Results' in xl.sheet_names else xl.sheet_names[0]
                if sheet == 'Summary' and len(xl.sheet_names) > 1:
                    sheet = xl.sheet_names[1]
                df = pd.read_excel(xl, sheet_name=sheet, dtype=str).fillna('')
                ft_cols = [c for c in df.columns if c.startswith('Full Text (PMC)')]
                if ft_cols:
                    df['_fulltext'] = df[ft_cols].apply(
                        lambda r: ' '.join(str(v) for v in r if v and str(v) not in ('','nan')), axis=1)
                articles = df.to_dict(orient='records')
            elif fname.endswith('.zip'):
                articles = [{'title': 'PDF ZIP file uploaded', 'abstract': 'PDF content available'}]
        except Exception as e:
            return jsonify({'error': f'Could not read file: {e}'}), 400

        # ── Build compact context — cap at 20 articles, 120 chars per abstract ──
        context_lines = [f"Dataset: {fname} ({len(articles)} total articles)\n"]
        for i, art in enumerate(articles[:20], 1):
            def _g(*keys):
                for k in keys:
                    v = art.get(k,'') or ''
                    if v and str(v).strip() not in ('','nan'): return str(v).strip()
                return ''
            title    = _g('Title','title')[:100]
            abstract = _g('Abstract','abstract')[:120]
            journal  = _g('Journal','journal')[:50]
            pub_date = _g('Publication Date','publication_date')
            country  = _g('Country','country')
            pub_type = _g('Publication Type','publication_type')
            context_lines.append(
                f"{i}. {title} | {journal} | {pub_date} | {country} | {pub_type}\n"
                f"   {abstract}\n"
            )

        if len(articles) > 20:
            context_lines.append(f"... and {len(articles)-20} more articles (showing first 20 for context)\n")

        article_context = '\n'.join(context_lines)

        # Keep total system prompt under 2000 chars
        system_prompt = (
            f"You are EpiLite Co-pilot, a biomedical research assistant. "
            f"The user uploaded {len(articles)} PubMed articles. "
            f"Answer questions based on this data:\n\n{article_context}\n\n"
            f"Be concise, specific, and cite article titles when relevant."
        )

        messages = [{'role': 'system', 'content': system_prompt}]
        for msg in history[-4:]:
            messages.append({'role': msg['role'], 'content': str(msg['content'])[:500]})
        messages.append({'role': 'user', 'content': question[:1000]})

        answer, err = call_azure_openai(messages, max_tokens=1200, temperature=0.3)
        if err:
            return jsonify({'error': err}), 500
        return jsonify({'answer': answer, 'articles_used': min(len(articles), 20)})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/extract_report', methods=['POST'])
@login_required
def extract_report():
    """Accept uploaded CSV/Excel/ZIP, start background extraction job."""
    import uuid as _uuid
    import zipfile as _zipfile
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'No file uploaded'}), 400
    fname = f.filename or ''
    articles = []
    try:
        if fname.endswith('.csv'):
            df = pd.read_csv(f, encoding='utf-8-sig', dtype=str).fillna('')
            articles = df.to_dict(orient='records')
        elif fname.endswith(('.xlsx', '.xls')):
            xl = pd.ExcelFile(f)
            sheet = 'All Results' if 'All Results' in xl.sheet_names else xl.sheet_names[0]
            if sheet == 'Summary' and len(xl.sheet_names) > 1:
                sheet = xl.sheet_names[1]
            df = pd.read_excel(xl, sheet_name=sheet, dtype=str).fillna('')
            ft_cols = [c for c in df.columns if c.startswith('Full Text (PMC)')]
            if ft_cols:
                df['full_text_merged'] = df[ft_cols].apply(
                    lambda r: ' '.join(str(v) for v in r if v and str(v) not in ('','nan')), axis=1)
            articles = df.to_dict(orient='records')
        elif fname.endswith('.zip'):
            zip_bytes = f.read()
            zf = _zipfile.ZipFile(io.BytesIO(zip_bytes))
            pdf_files = [n for n in zf.namelist() if n.lower().endswith('.pdf')]
            if not pdf_files:
                return jsonify({'error': 'No PDF files found in ZIP'}), 400
            for pdf_name in pdf_files[:100]:
                try:
                    pdf_bytes = zf.read(pdf_name)
                    text = ''
                    if PDF_SUPPORT:
                        import pdfplumber
                        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
                            text = '\n\n'.join(page.extract_text() or '' for page in pdf.pages[:30])
                    clean_name = pdf_name.replace('.pdf','').replace('_',' ')
                    pmid = ''
                    if clean_name.startswith('PMID '):
                        parts = clean_name.split(' ', 2)
                        pmid = parts[1] if len(parts)>1 else ''
                        title_hint = parts[2] if len(parts)>2 else clean_name
                    else:
                        title_hint = clean_name
                    articles.append({
                        'pmid': pmid, 'title': title_hint[:120],
                        'abstract': text[:500], 'full_text': text,
                        'url': f'https://pubmed.ncbi.nlm.nih.gov/{pmid}/' if pmid else '',
                        '_source': 'pdf', '_filename': pdf_name,
                    })
                except Exception as e:
                    print(f"[extract] PDF read error {pdf_name}: {e}")
        else:
            return jsonify({'error': 'Upload CSV, Excel (.xlsx), or PDF ZIP'}), 400
    except Exception as e:
        return jsonify({'error': f'Could not read file: {e}'}), 400

    if not articles:
        return jsonify({'error': 'No articles found in file'}), 400
    articles = articles[:100]

    job_id = _uuid.uuid4().hex[:12]
    _extract_jobs[job_id] = {
        'status': 'Starting...', 'done': False, 'current': 0,
        'total': len(articles), 'excel_bytes': None, 'error': None, 'cancelled': False,
    }
    t = threading.Thread(target=_run_extract_job, args=(job_id, articles), daemon=True)
    t.start()
    return jsonify({'job_id': job_id, 'total': len(articles)})


@app.route('/api/extract_progress/<job_id>')
@login_required
def extract_progress(job_id):
    job = _extract_jobs.get(job_id)
    if not job:
        return jsonify({'error': 'Job not found'}), 404
    return jsonify({
        'status': job['status'], 'done': job['done'],
        'current': job['current'], 'total': job['total'], 'error': job['error'],
    })


@app.route('/api/extract_download/<job_id>')
@login_required
def extract_download(job_id):
    job = _extract_jobs.get(job_id)
    if not job or not job['done'] or not job['excel_bytes']:
        return "Job not ready or not found", 404
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(
        io.BytesIO(job['excel_bytes']),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'AE_Extraction_{ts}.xlsx',
    )


@app.route('/api/extract_article', methods=['POST'])
@login_required
def extract_article():
    """Extract 51-column AE report from a single attached article file."""
    import uuid as _uuid
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'No file attached'}), 400
    fname = f.filename or ''
    text  = ''
    try:
        if fname.lower().endswith('.pdf'):
            if PDF_SUPPORT:
                import pdfplumber
                with pdfplumber.open(io.BytesIO(f.read())) as pdf:
                    text = '\n\n'.join(page.extract_text() or '' for page in pdf.pages[:40])
            else:
                return jsonify({'error': 'PDF support not available. Use TXT or Excel.'}), 400
        elif fname.lower().endswith('.txt'):
            text = f.read().decode('utf-8', errors='ignore')
        elif fname.lower().endswith(('.xlsx','.xls')):
            df = pd.read_excel(f, dtype=str).fillna('')
            text = df.to_string(index=False)[:8000]
        elif fname.lower().endswith('.csv'):
            df = pd.read_csv(f, encoding='utf-8-sig', dtype=str).fillna('')
            text = df.to_string(index=False)[:8000]
        else:
            return jsonify({'error': 'Unsupported file type. Use PDF, TXT, Excel, or CSV.'}), 400
    except Exception as e:
        return jsonify({'error': f'Could not read file: {e}'}), 400
    if not text.strip():
        return jsonify({'error': 'No text could be extracted from the file.'}), 400
    article = {
        'title': fname.replace('.pdf','').replace('.txt','').replace('_',' ')[:120],
        'abstract': text[:1000], 'full_text': text,
        'url': '', 'authors': '', 'journal': '',
        'publication_date': '', 'publication_type': '',
        'country': '', 'pmid': '',
        '_source': 'attached_article', '_filename': fname,
    }
    job_id = _uuid.uuid4().hex[:12]
    _extract_jobs[job_id] = {
        'status': f'Extracting from {fname}...', 'done': False,
        'current': 0, 'total': 1, 'excel_bytes': None, 'error': None, 'cancelled': False,
    }
    t = threading.Thread(target=_run_extract_job, args=(job_id, [article]), daemon=True)
    t.start()
    return jsonify({'job_id': job_id, 'total': 1, 'filename': fname})


@app.route('/api/extract_article_template', methods=['POST'])
@login_required
def extract_article_template():
    """
    Extract clinical data from a PDF/TXT article and populate
    the user's own Excel template columns (or default 51 columns).

    Accepts:
      - article: PDF or TXT file (the article to extract from)
      - template: Excel file with column headers in row 1 (optional)
      - question: user's message
    """
    import uuid as _uuid

    article_file  = request.files.get('article')
    template_file = request.files.get('template')

    if not article_file:
        return jsonify({'error': 'Please attach a PDF or TXT article file'}), 400

    # ── Read article content ──────────────────────────────────────────
    fname = article_file.filename or ''
    text  = ''
    try:
        if fname.lower().endswith('.pdf'):
            if PDF_SUPPORT:
                import pdfplumber
                with pdfplumber.open(io.BytesIO(article_file.read())) as pdf:
                    pages = []
                    for page in pdf.pages[:40]:
                        pt = page.extract_text() or ''
                        pages.append(pt)
                    text = '\n\n'.join(pages)
            else:
                return jsonify({'error': 'PDF support not available on server. Upload TXT instead.'}), 400
        elif fname.lower().endswith('.txt'):
            text = article_file.read().decode('utf-8', errors='ignore')
        else:
            return jsonify({'error': 'Please attach a PDF or TXT article'}), 400
    except Exception as e:
        return jsonify({'error': f'Could not read article: {e}'}), 400

    if not text.strip():
        return jsonify({'error': 'No text could be extracted from the file'}), 400

    # ── Read template columns ─────────────────────────────────────────
    template_columns = []
    if template_file:
        try:
            tmpl_name = template_file.filename or ''
            if tmpl_name.lower().endswith(('.xlsx', '.xls')):
                df_tmpl = pd.read_excel(io.BytesIO(template_file.read()), header=0, nrows=0)
                template_columns = [c for c in df_tmpl.columns if c and str(c).strip() not in ('', 'nan', 'Unnamed')]
            print(f"[extract_article] Template columns ({len(template_columns)}): {template_columns[:10]}")
        except Exception as e:
            print(f"[extract_article] Template read error: {e}")
            template_columns = []

    # ── Build article dict for _run_extract_job ───────────────────────
    article = {
        'title':            fname.replace('.pdf','').replace('.txt','').replace('_',' ')[:120],
        'abstract':         text[:800],
        'full_text':        text,
        'full_text_merged': text,
        'url':              '',
        'authors':          '',
        'journal':          '',
        'publication_date': '',
        'publication_type': '',
        'country':          '',
        'pmid':             '',
        '_source':          'attached_pdf',
        '_filename':        fname,
        '_template_cols':   template_columns,
    }

    job_id = _uuid.uuid4().hex[:12]
    _extract_jobs[job_id] = {
        'status':         f'Extracting from {fname}...',
        'done':           False,
        'current':        0,
        'total':          1,
        'excel_bytes':    None,
        'error':          None,
        'cancelled':      False,
        'template_cols':  template_columns,
    }

    t = threading.Thread(
        target=_run_extract_with_template,
        args=(job_id, article, template_columns),
        daemon=True
    )
    t.start()

    return jsonify({
        'job_id':          job_id,
        'total':           1,
        'filename':        fname,
        'template_cols':   len(template_columns),
    })


def _run_extract_with_template(job_id: str, article: dict, template_columns: list):
    """
    Extract clinical data from one article and write to Excel.
    If template_columns provided, use those as column headers.
    Otherwise use the default EXTRACT_COLUMNS.
    """
    job      = _extract_jobs[job_id]
    if not os.environ.get('AZURE_OPENAI_KEY', ''):
        job['error'] = 'AZURE_OPENAI_KEY not set in Render environment'
        job['done']  = True
        return

    job['current'] = 1
    job['status']  = f"Extracting from {article.get('_filename', 'article')}..."

    text     = article.get('full_text', '') or article.get('abstract', '')
    fname    = article.get('_filename', 'article')

    # Use template columns if provided, otherwise default
    use_cols = template_columns if template_columns else [c for c in EXTRACT_COLUMNS if c]

    # Truncate content smartly
    if len(text) > 3000:
        content = text[:2000] + '\n...\n' + text[-1000:]
    else:
        content = text

    # Build dynamic prompt based on template columns
    cols_str = '\n'.join(f'  "{c}": ""' for c in use_cols)

    system_msg = ('You are a clinical research data extractor. '
                  'Extract data from the article and return ONLY valid JSON. '
                  'No markdown, no explanation.')

    user_msg = f"""Extract clinical data from this article and fill ALL the columns listed below.

Article filename: {fname}

Content:
{content}

Return this exact JSON with all fields filled from the article content:
{{
  "rows": [
    {{
{cols_str}
    }}
  ]
}}

- Fill every field you can find in the content
- Use empty string "" only if truly not mentioned
- Create multiple rows if multiple adverse events are reported
- For AE grade columns: extract percentages and denominators from results tables"""

    try:
        messages = [
            {'role': 'system', 'content': system_msg},
            {'role': 'user',   'content': user_msg},
        ]
        raw, api_err = call_azure_openai(messages, max_tokens=2000, temperature=0)
        print(f"[extract_tmpl] Azure response: {(raw or '')[:400]}")
        rows = []

        if raw is not None:
            print(f"[extract_tmpl] Raw (first 400): {raw[:400]}")
            raw = re.sub(r'^```[a-z]*\n?', '', raw.strip())
            raw = re.sub(r'\n?```$', '', raw.strip())
            jm  = re.search(r'\{.*\}', raw, re.DOTALL)
            if jm:
                extracted = json.loads(jm.group())
                rows = extracted.get('rows', [])
                print(f"[extract_tmpl] Got {len(rows)} rows")
        else:
            print(f"[extract_tmpl] API error: {api_err}")

        if not rows:
            # Fallback row with just the filename
            rows = [{c: '' for c in use_cols}]
            rows[0][use_cols[0]] = fname if use_cols else ''

    except Exception as e:
        print(f"[extract_tmpl] Exception: {e}")
        rows = [{c: '' for c in use_cols}]

    # ── Build Excel with template columns ─────────────────────────────
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = 'AE Extraction'

        hdr_fill = PatternFill('solid', start_color='1A365D')
        hdr_font = Font(name='Arial', bold=True, color='FFFFFF', size=9)
        hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Write headers
        for ci, col in enumerate(use_cols, 1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.fill  = hdr_fill
            cell.font  = hdr_font
            cell.alignment = hdr_align
            ws.column_dimensions[cell.column_letter].width = max(15, min(40, len(col) + 5))

        ws.row_dimensions[1].height = 45
        ws.freeze_panes = 'A2'

        # Write data rows
        for ri, row_data in enumerate(rows, start=2):
            for ci, col in enumerate(use_cols, 1):
                val  = row_data.get(col, '') or ''
                cell = ws.cell(row=ri, column=ci, value=str(val) if val else '')
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                cell.font = Font(name='Arial', size=9)
                if ri % 2 == 0:
                    cell.fill = PatternFill('solid', start_color='EBF4FF')

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        job['excel_bytes'] = buf.getvalue()
        job['done']   = True
        job['status'] = f'Complete -- {len(rows)} row(s) extracted from {fname}'
        print(f"[extract_tmpl] Excel built OK: {len(rows)} rows, {len(use_cols)} columns")

    except Exception as e:
        job['error'] = f'Excel build failed: {e}'
        job['done']  = True
        import traceback; traceback.print_exc()



@login_required
def test_extraction():
    """Debug: test extraction on first article, return raw Groq response."""
    f = request.files.get('file')
    if not f: return jsonify({'error': 'No file'}), 400
    fname = f.filename or ''
    try:
        if fname.endswith('.csv'):
            df = pd.read_csv(f, encoding='utf-8-sig', dtype=str).fillna('')
        elif fname.endswith(('.xlsx','.xls')):
            xl = pd.ExcelFile(f)
            sheet = 'All Results' if 'All Results' in xl.sheet_names else xl.sheet_names[0]
            df = pd.read_excel(xl, sheet_name=sheet, dtype=str).fillna('')
        else:
            return jsonify({'error': 'Upload CSV or Excel'}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 400
    if df.empty: return jsonify({'error': 'Empty file'}), 400
    art = df.iloc[0].to_dict()
    def _g(*keys):
        for k in keys:
            v = art.get(k,'') or ''
            if v and str(v).strip() not in ('','nan'): return str(v).strip()
        return ''
    title    = _g('Title','title')
    abstract = _g('Abstract','abstract')
    authors  = _g('Authors','authors')
    country  = _g('Country','country')
    url      = _g('PubMed URL','url')
    pub_date = _g('Publication Date','publication_date')
    journal  = _g('Journal','journal')
    content  = abstract[:1500]
    year     = re.search(r'\b(19|20)\d{2}\b', pub_date).group() if re.search(r'\b(19|20)\d{2}\b', pub_date) else ''
    first_a  = authors.split(',')[0].strip().split()[-1] if authors else ''
    author_year = (first_a + ' et al., ' + year).strip(' ,') if first_a else authors[:40]
    if not os.environ.get('AZURE_OPENAI_KEY',''): return jsonify({'error': 'AZURE_OPENAI_KEY not set', 'columns': list(df.columns[:10])}), 500
    try:
        AZURE_ENDPOINT   = os.environ.get('AZURE_OPENAI_ENDPOINT','').rstrip('/')
        AZURE_DEPLOYMENT = os.environ.get('AZURE_OPENAI_DEPLOYMENT','gpt-4o-mini')
        AZURE_KEY        = os.environ.get('AZURE_OPENAI_KEY','')
        resp = http_requests.post(
            f"{AZURE_ENDPOINT}/openai/deployments/{AZURE_DEPLOYMENT}/chat/completions?api-version=2024-02-01",
            headers={'Content-Type':'application/json','api-key':AZURE_KEY},
            json={'max_tokens':500,'temperature':0,
                  'messages':[
                      {'role':'system','content':'Return ONLY valid JSON. No markdown.'},
                      {'role':'user','content':f'Title: {title}\nAbstract: {content}\nReturn: {{"rows":[{{"author_year":"{author_year}","country":"{country}","study_title":"{title[:50].replace(chr(34),chr(39))}","cancer_name":"","treatment":"","url":"{url}"}}]}}'  }]},
            timeout=30,
        )
        raw = resp.json().get('choices',[{}])[0].get('message',{}).get('content','') if resp.status_code==200 else resp.text
        return jsonify({'api_status':resp.status_code,'title':title,'abstract_length':len(abstract),'columns':list(df.columns[:10]),'azure_raw':raw[:800],'parsed_ok':'rows' in raw})
    except Exception as e:
        return jsonify({'error':str(e)}), 500


if __name__ == '__main__':
    print("Starting EpiLite AI Platform...")
    app.run(debug=True, host='0.0.0.0', port=5000)
